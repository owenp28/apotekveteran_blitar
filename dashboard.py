import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import io
import re
from datetime import date, datetime, timezone, timedelta
from pathlib import Path
import json
import time
import base64
import sqlite3
import zipfile
from io import BytesIO
from urllib.request import Request, urlopen
from openpyxl import load_workbook, Workbook

# [SAFE IMPORT SQLALCHEMY]
HAS_SQLALCHEMY = False
try:
    import sqlalchemy as sa
    HAS_SQLALCHEMY = True
except ImportError:
    HAS_SQLALCHEMY = False

# Konfigurasi Halaman Streamlit
st.set_page_config(page_title="Apotek Veteran Blitar", layout="wide", page_icon="💊")

# ── SETUP KONEKSI DATABASE (HYBRID: SUPABASE / POSTGRESQL / BUILT-IN SQLITE) ──
SQLITE_DB_PATH = Path(__file__).parent / "apotek_veteran.db"

def get_db_engine():
    if HAS_SQLALCHEMY:
        try:
            db_url = st.secrets["DB_URL"]
            return sa.create_engine(db_url)
        except Exception:
            return sa.create_engine(f"sqlite:///{SQLITE_DB_PATH}")
    return None

engine = get_db_engine()

# Helper Eksekusi Query Database yang Kompatibel (SQLAlchemy & Native SQLite3)
def db_read_table(table_name):
    try:
        if engine is not None:
            return pd.read_sql(table_name, engine)
        else:
            with sqlite3.connect(SQLITE_DB_PATH) as conn:
                return pd.read_sql(f"SELECT * FROM {table_name}", conn)
    except Exception:
        return pd.DataFrame()

def db_write_table(df, table_name):
    try:
        if engine is not None:
            df.to_sql(table_name, engine, if_exists="replace", index=False)
        else:
            with sqlite3.connect(SQLITE_DB_PATH) as conn:
                df.to_sql(table_name, conn, if_exists="replace", index=False)
        return True
    except Exception as e:
        st.error(f"Gagal menyimpan ke Database: {e}")
        return False

# ── CSS Custom ERP ────────────────────────────────────────────────────────────
st.markdown(
    """
    <style>
    * { margin: 0; padding: 0; box-sizing: border-box; }
    body { background: #1a1a2e; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; color: #e0e0e0; }
    [data-testid="stSidebar"] > div:first-child { padding-top: 2rem !important; }
    .block-container { padding-top: 3rem !important; padding-bottom: 2rem !important; margin-top: 0rem !important; padding-left: 20px !important; padding-right: 20px !important; }
    .app-header { text-align: center; margin-bottom: 30px; padding: 20px; background: linear-gradient(135deg, #16213e 0%, #0f3460 100%); border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.3); }
    .app-title { font-size: 42px; font-weight: 700; color: #e94560; margin-bottom: 10px; text-shadow: 2px 2px 4px rgba(0,0,0,0.3); }
    .app-subtitle { font-size: 16px; color: #a0a0a0; font-weight: 400; }
    .form-container { background: #16213e; border-radius: 12px; padding: 20px; margin-bottom: 20px; border: 1px solid #0f3460; box-shadow: 0 4px 15px rgba(0,0,0,0.2); }
    .form-section-title { font-size: 18px; font-weight: 600; color: #e94560; margin-bottom: 15px; padding-bottom: 10px; border-bottom: 2px solid #e94560; display: flex; align-items: center; gap: 10px; }
    .btn-custom { padding: 12px 24px; border-radius: 6px; font-size: 14px; font-weight: 600; cursor: pointer; transition: all 0.3s ease; display: inline-flex; align-items: center; gap: 8px; border: none; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; }
    [data-testid="stFileUploadDropzone"] { padding: 5px !important; min-height: 42px !important; border-radius: 6px !important; }
    [data-testid="stFileUploadDropzone"] > div { padding-top: 0 !important; padding-bottom: 0 !important; }
    .table-container { background: #16213e; border-radius: 12px; padding: 20px; margin-bottom: 20px; border: 1px solid #0f3460; box-shadow: 0 4px 15px rgba(0,0,0,0.2); }
    .stDataFrame { background: #1a1a2e; border-radius: 8px; overflow: hidden; }
    .stDataFrame th { background: #0f3460; color: #e0e0e0; font-weight: 600; font-size: 13px; padding: 10px; }
    .stDataFrame td { color: #e0e0e0; font-size: 13px; padding: 8px; }
    .stDataFrame tr:hover { background: #1f3a5e; }
    .filter-label { font-size: 12px; font-weight: bold; color: #a0a0a0; margin-bottom: 4px; display: block; }
    </style>
    """,
    unsafe_allow_html=True
)

# ── FUNGSI WAKTU LOKAL (WIB) ──────────────────────────────────────────────────
def get_wib_time():
    utc_now = datetime.now(timezone.utc)
    wib_now = utc_now + timedelta(hours=7)
    return wib_now.replace(tzinfo=None)

# ── KONFIGURASI KOLOM STANDAR ─────────────────────────────────────────────────
DEFAULT_LINK_ONEDRIVE = "https://1drv.ms/x/c/2b91c5c1ac3eaa9f/IQBzkm7nxPNlRI4V4fKaVYERASx-hzJiaBEWDdCFPu79k3w?e=V5jQMP"

INVENTORY_SHEETS = ["PCS", "SACHET", "BOTOL", "TAB", "BOX", "STRIP"]
INVENTORY_COLUMNS = [
    "Worksheet", "Nama produk", "Satuan", "Tanggal", "Nomor Faktur", "Nomor Batch", 
    "PBF", "Tanggal Kadaluwarsa", "Stok Masuk", "Stok Keluar", "Stok Sisa", 
    "Harga 1", "Harga 2", "Keterangan"
]

KOLOM_WAJIB = [
    "Tanggal", "Nomor Faktur", "Nama Obat", "Kategori", "Satuan", "Nomor Batch",
    "Stok Masuk", "Stok Keluar", "Stok Akhir", "Harga Satuan (Rp)", "Total Nilai (Rp)", 
    "Tanggal Kadaluarsa", "Keterangan", "Petugas"
]
RETUR_HISTORY_COLUMNS = ["Nomor Faktur", "Tanggal Retur", "Jumlah Item", "Total Nilai Retur", "Tanggal Disimpan"]
SHIFT_COLUMNS = [
    "Waktu Buka", "Waktu Tutup", "Shift", "Pendaftar Shift", "Kasir Bergabung", "Saldo Awal", 
    "Hasil Penjualan", "Piutang", "Pendapatan Jurnal", "Total Pendapatan",
    "Retur Penjualan", "Pengeluaran Jurnal", "Total Pengeluaran",
    "Saldo Akhir Sistem", "Fisik Kasir Aktual", "Selisih", "Diserahkan Ke", "Catatan"
]

# ── PERSISTENSI SHIFT LOGIC (DATABASE) ────────────────────────────────────────
def load_active_shift():
    try:
        df = db_read_table("active_shift_state")
        if not df.empty:
            data = df.iloc[0].to_dict()
            data["joined_users"] = json.loads(data["joined_users"]) if isinstance(data.get("joined_users"), str) else []
            return data
    except Exception: pass
    return None

def save_active_shift(data):
    df_save = data.copy()
    df_save["joined_users"] = json.dumps(df_save.get("joined_users", []))
    df = pd.DataFrame([df_save])
    db_write_table(df, "active_shift_state")

def clear_active_shift():
    try:
        if engine is not None:
            with engine.begin() as conn:
                conn.execute(sa.text("DROP TABLE IF EXISTS active_shift_state"))
        else:
            with sqlite3.connect(SQLITE_DB_PATH) as conn:
                conn.execute("DROP TABLE IF EXISTS active_shift_state")
    except Exception: pass

def get_auto_shift_name():
    hour = get_wib_time().hour
    if 4 <= hour < 10: return "Pagi"
    elif 10 <= hour < 15: return "Siang"
    elif 15 <= hour < 18: return "Sore"
    else: return "Malam"

# ── FUNGSI DATA I/O (DATABASE) ────────────────────────────────────────────────
def load_data():
    df = db_read_table("history_transaksi")
    if df.empty:
        return pd.DataFrame(columns=KOLOM_WAJIB)
    for col in ["Tanggal", "Tanggal Kadaluarsa"]:
        if col in df.columns: df[col] = pd.to_datetime(df[col], errors="coerce")
    return df

def save_data(df):
    db_write_table(df, "history_transaksi")

def load_retur_history():
    df = db_read_table("history_retur")
    if df.empty:
        return pd.DataFrame(columns=RETUR_HISTORY_COLUMNS)
    for col in ["Tanggal Retur", "Tanggal Disimpan"]:
        if col in df.columns: df[col] = pd.to_datetime(df[col], errors="coerce")
    return df

def save_retur_history(df):
    db_write_table(df, "history_retur")

def load_shift_log():
    df = db_read_table("log_shift")
    if df.empty:
        return pd.DataFrame(columns=SHIFT_COLUMNS)
    return df

def save_shift_log(df):
    db_write_table(df, "log_shift")

def get_backup_zip():
    mem_zip = io.BytesIO()
    with zipfile.ZipFile(mem_zip, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for t_name in ["inventory_master", "history_transaksi", "history_retur", "log_shift"]:
            df_t = db_read_table(t_name)
            if not df_t.empty:
                zf.writestr(f"{t_name}.csv", df_t.to_csv(index=False))
    return mem_zip.getvalue()

def format_rupiah(val):
    try: return f"Rp {int(val):,}".replace(",", ".")
    except: return val

def parse_excel_date(val):
    if pd.isna(val): return pd.NaT
    val_str = str(val).strip()
    if val_str in ["", "-", "nan", "None", "NaT", "0", "0.0"]: return pd.NaT
    if isinstance(val, (datetime, date)):
        d = val.date() if isinstance(val, datetime) else val
        return pd.Timestamp(d) if d.year > 1970 else pd.NaT
    try:
        f_val = float(val)
        if f_val > 10000:
            d = (pd.Timestamp('1899-12-30') + pd.Timedelta(days=f_val)).date()
            return pd.Timestamp(d) if d.year > 1970 else pd.NaT
        return pd.NaT
    except Exception: pass
    try:
        d = pd.to_datetime(val)
        return pd.Timestamp(d) if d.year > 1970 else pd.NaT
    except Exception: return pd.NaT

def normalize_inventory_df(df):
    df = df.copy()
    renamed = {}
    for kolom in df.columns:
        nama_kolom = str(kolom).strip()
        if nama_kolom.lower() in ["nama obat", "nama produk"]: renamed[kolom] = "Nama produk"
        elif nama_kolom.lower() == "pbf ": renamed[kolom] = "PBF"
        elif nama_kolom.lower() == "keterangan ": renamed[kolom] = "Keterangan"
    if renamed: df = df.rename(columns=renamed)
    
    # Amankan filter khusus data Master Dataset
    cols_check = [c for c in INVENTORY_COLUMNS if c != "Worksheet"]
    for kolom in cols_check:
        if kolom not in df.columns: df[kolom] = None

    for kolom in ["Nama produk", "Satuan", "Nomor Faktur", "Nomor Batch", "PBF", "Keterangan"]:
        if kolom in df.columns: df[kolom] = df[kolom].astype("string")

    for kolom in ["Stok Masuk", "Stok Keluar", "Stok Sisa", "Harga 1", "Harga 2"]:
        if kolom in df.columns: df[kolom] = pd.to_numeric(df[kolom], errors="coerce")

    for col in ["Tanggal", "Tanggal Kadaluwarsa"]:
        if col in df.columns: df[col] = df[col].apply(parse_excel_date)
        
    final_cols = cols_check.copy()
    if "Worksheet" in df.columns:
        final_cols.insert(0, "Worksheet")
    
    return df[final_cols]

def prepare_sheet_for_editor(df):
    df = normalize_inventory_df(df)
    for kolom in ["Nomor Faktur", "Nomor Batch", "PBF", "Keterangan", "Nama produk", "Satuan", "Worksheet"]:
        if kolom in df.columns: df[kolom] = df[kolom].astype("string")
    return df

# [PERBAIKAN] - Pendeteksi Cerdas Baris Judul
def _find_inventory_header_row(rows):
    known_headers = {"nama produk", "nama obat", "satuan", "tanggal", "nomor faktur", "nomor batch", "pbf", "tanggal kadaluarsa", "stok masuk", "stok keluar", "stok sisa", "harga 1", "harga 2", "keterangan"}
    for index, row in enumerate(rows):
        cleaned = [str(cell).strip().lower() if cell is not None else "" for cell in row]
        score = sum(1 for cell in cleaned if cell in known_headers)
        if score >= 3: return index, list(row)
    return 0, list(rows[0]) if rows else []

# [PERBAIKAN] - Membaca File Excel Langsung Lewat OpenPyxl Agar Header Terdeteksi
def load_inventory_sheet_dataframe(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return pd.DataFrame(columns=[c for c in INVENTORY_COLUMNS if c != "Worksheet"])
    header_index, raw_header = _find_inventory_header_row(rows)
    header = [str(cell).strip() if cell is not None else "" for cell in raw_header]
    data_rows = rows[header_index + 1:]
    if not data_rows: return pd.DataFrame(columns=[c for c in INVENTORY_COLUMNS if c != "Worksheet"])
    
    # Amankan panjang kolom jika tidak sama
    data_rows = [tuple(row[:len(header)]) for row in data_rows]
    df = pd.DataFrame(data_rows, columns=header)
    return normalize_inventory_df(df)

def load_inventory_workbook():
    df_all = db_read_table("inventory_master")
    if df_all.empty:
        return {}
    workbook_data = {}
    if "Worksheet" not in df_all.columns: df_all["Worksheet"] = "TAB"
    for sheet in df_all["Worksheet"].unique():
        df_sheet = df_all[df_all["Worksheet"] == sheet].drop(columns=["Worksheet"], errors='ignore')
        workbook_data[sheet] = df_sheet
    return workbook_data

def save_inventory_workbook(workbook_data):
    frames = []
    for sheet_name, df_sheet in workbook_data.items():
        if df_sheet is None or df_sheet.empty: 
            df_sheet = pd.DataFrame(columns=[c for c in INVENTORY_COLUMNS if c != "Worksheet"])
        df_copy = df_sheet.copy()
        df_copy["Worksheet"] = sheet_name
        frames.append(df_copy)
        
    if frames:
        combined = pd.concat(frames, ignore_index=True)
        # Hapus baris yang kosong atau tidak valid namanya
        combined = combined.dropna(subset=["Nama produk"], how="all")
        for col in combined.columns:
            combined[col] = combined[col].apply(lambda v: str(v) if isinstance(v, (list, tuple, dict, set)) else v)
        return db_write_table(combined, "inventory_master")
    return False

def build_inventory_print_dataframe():
    workbook_data = st.session_state.get("inventory_data_cache")
    if not workbook_data:
        workbook_data = load_inventory_workbook()
        st.session_state.inventory_data_cache = workbook_data
    if not workbook_data: return None
    
    frames = []
    for sheet_name, df_sheet in workbook_data.items():
        df_copy = prepare_sheet_for_editor(df_sheet.copy())
        df_copy["Worksheet"] = sheet_name
        frames.append(df_copy)
        
    if not frames: return pd.DataFrame(columns=INVENTORY_COLUMNS)
    return pd.concat(frames, ignore_index=True)

def get_available_sheets():
    cache = st.session_state.get("inventory_data_cache", {})
    if cache: return list(cache.keys())
    return INVENTORY_SHEETS

# ── LOGIKA BACKEND PROSES OPNAME ──
def do_opname_processing(edited_opname_df):
    workbook_data = st.session_state.inventory_data_cache
    df_history = load_data()
    new_history_rows = []
    
    for idx, row in edited_opname_df.iterrows():
        lokasi = row.get("Worksheet", None)
        if pd.isna(lokasi) or not str(lokasi).strip() or str(lokasi).strip() not in workbook_data:
            continue
            
        lokasi_str = str(lokasi).strip()
        nama = str(row.get("Nama Obat", "")).strip()
        batch = str(row.get("No. Batch", "")).strip()
        
        try: nyata = float(row.get("Stok Nyata Terkecil", 0.0))
        except: nyata = 0.0
        
        try: sistem = float(row.get("Stok Satuan Terkecil", 0.0))
        except: sistem = 0.0
        
        try: stok_exp = float(row.get("Stok Expired Terkecil", 0.0))
        except: stok_exp = 0.0
        
        if nyata > 0 or stok_exp > 0:
            ws_target = prepare_sheet_for_editor(workbook_data[lokasi_str].copy())
            mask_target = (ws_target["Nama produk"].astype(str).str.strip() == nama) & (ws_target["Nomor Batch"].astype(str).str.strip() == batch)
            
            if mask_target.any():
                target_idx = ws_target[mask_target].index[-1]
                ws_target.loc[target_idx, "Stok Sisa"] = nyata
                diff = nyata - sistem
                
                if diff != 0:
                    satuan_obat = str(ws_target.loc[target_idx, "Satuan"]) if pd.notna(ws_target.loc[target_idx, "Satuan"]) else ""
                    new_history_rows.append({
                        "Tanggal": get_wib_time().strftime("%Y-%m-%d %H:%M:%S"), 
                        "Nomor Faktur": "OPNAME", 
                        "Nama Obat": nama, 
                        "Kategori": lokasi_str, 
                        "Satuan": satuan_obat, 
                        "Nomor Batch": batch,
                        "Stok Masuk": diff if diff > 0 else 0, 
                        "Stok Keluar": abs(diff) if diff < 0 else 0, 
                        "Stok Akhir": nyata, 
                        "Harga Satuan (Rp)": 0.0, 
                        "Total Nilai (Rp)": 0.0, 
                        "Tanggal Kadaluarsa": pd.NaT, 
                        "Keterangan": "Proses stokopname", 
                        "Petugas": USERS.get(st.session_state.username, {}).get("name", "Sistem")
                    })
                
                if diff > 0: 
                    stok_masuk_lama = float(ws_target.loc[target_idx, "Stok Masuk"]) if pd.notna(ws_target.loc[target_idx, "Stok Masuk"]) else 0.0
                    ws_target.loc[target_idx, "Stok Masuk"] = stok_masuk_lama + diff
                elif diff < 0: 
                    stok_keluar_lama = float(ws_target.loc[target_idx, "Stok Keluar"]) if pd.notna(ws_target.loc[target_idx, "Stok Keluar"]) else 0.0
                    ws_target.loc[target_idx, "Stok Keluar"] = stok_keluar_lama + abs(diff)
                workbook_data[lokasi_str] = ws_target
                
    if new_history_rows:
        df_history = pd.concat([df_history, pd.DataFrame(new_history_rows)], ignore_index=True)
        save_data(df_history)

    save_inventory_workbook(workbook_data)
    st.session_state.inventory_data_cache = workbook_data

# ── STOK OPNAME MODAL & LOGIN ─────────────────────────────────────────────────
@st.dialog("Pilih Obat", width="large")
def modal_pilih_obat(df_source, initial_search=""):
    st.markdown("<style>.stDialog > div { padding: 15px; }</style>", unsafe_allow_html=True)
    col_c1, col_c2, col_c3 = st.columns([3, 3, 4])
    with col_c1: tampil_habis = st.checkbox("Tampilkan No. Batch yang sudah habis", value=False, key="modal_chk_habis")
    with col_c2: tampil_stok = st.checkbox("Tampilkan Stok", value=True, key="modal_chk_stok")
    with col_c3: urut_ed = st.checkbox("Urutkan berdasarkan ED terdekat (FEFO)", value=True, key="modal_chk_fefo")

    st.button("🔄 Refresh", key="btn_refresh_modal")
    st.markdown("---")

    df_modal = df_source.copy()
    df_modal["Stok Sisa"] = pd.to_numeric(df_modal["Stok Sisa"], errors="coerce").fillna(0)
    
    if not tampil_habis: df_modal = df_modal[df_modal["Stok Sisa"] > 0]
    if urut_ed:
        df_modal["Tanggal Kadaluwarsa"] = pd.to_datetime(df_modal["Tanggal Kadaluwarsa"], errors="coerce")
        df_modal = df_modal.sort_values(by="Tanggal Kadaluwarsa", na_position="last")

    df_modal = df_modal.reset_index(drop=True)
    df_modal["No."] = range(1, len(df_modal) + 1)
    df_modal["Kode Obat"] = ["OBT" + str(1000 + i) for i in range(len(df_modal))]
    df_modal["Pilih"] = False
    
    if tampil_stok: df_modal["Stok Terkecil"] = df_modal.apply(lambda r: f"{r['Stok Sisa']:.2f} {str(r['Satuan']) if pd.notna(r['Satuan']) else str(r['Worksheet'])}", axis=1)
    else: df_modal["Stok Terkecil"] = "-"

    df_modal["Golongan"] = "-"
    df_modal["Kategori"] = "-"
    df_modal["Status Stok"] = df_modal["Stok Sisa"].apply(lambda x: "Stok Tersedia" if x > 0 else "Stok Habis")

    f0, f1, f2, f3, f4, f5, f6, f7 = st.columns([1, 1.5, 3.5, 1.5, 1.5, 1.5, 1.5, 1.5])
    with f1: f_kode = st.text_input("Kode", placeholder="Kode Obat", label_visibility="collapsed", key="f_kode")
    with f2: f_nama = st.text_input("Nama", value=initial_search, placeholder="Nama Obat", label_visibility="collapsed", key="f_nama")
    with f4: f_gol = st.selectbox("Golongan", ["Golongan", "-"], label_visibility="collapsed", key="f_gol")
    with f5: f_kat = st.selectbox("Kategori", ["Kategori", "-"], label_visibility="collapsed", key="f_kat")
    with f6: 
        satuan_opts = ["Semua Satuan"] + sorted(list(df_modal["Satuan"].dropna().astype(str).unique()))
        f_sat = st.selectbox("Satuan", satuan_opts, label_visibility="collapsed", key="f_sat")
    with f7: f_status = st.selectbox("Status", ["Status Stok", "Stok Tersedia", "Stok Habis"], label_visibility="collapsed", key="f_status")

    if f_kode: df_modal = df_modal[df_modal["Kode Obat"].str.contains(f_kode, case=False, na=False)]
    if f_nama: df_modal = df_modal[df_modal["Nama produk"].str.contains(f_nama, case=False, na=False)]
    if f_gol != "Golongan": df_modal = df_modal[df_modal["Golongan"] == f_gol]
    if f_kat != "Kategori": df_modal = df_modal[df_modal["Kategori"] == f_kat]
    if f_sat != "Semua Satuan": df_modal = df_modal[df_modal["Satuan"].astype(str) == f_sat]
    if f_status != "Status Stok": df_modal = df_modal[df_modal["Status Stok"] == f_status]

    cols = ["No.", "Pilih", "Kode Obat", "Nama produk", "Stok Terkecil", "Golongan", "Kategori", "Satuan", "Status Stok", "Worksheet"]
    df_display = df_modal[cols]

    st.caption(f"Menampilkan 1-{len(df_display)} dari {len(df_display)} data")

    edited_modal = st.data_editor(
        df_display,
        hide_index=True,
        use_container_width=True,
        column_config={
            "No.": st.column_config.NumberColumn("No.", width="small"),
            "Pilih": st.column_config.CheckboxColumn("Pilih", default=False),
            "Kode Obat": st.column_config.TextColumn("Kode Obat", width="medium"),
            "Nama produk": st.column_config.TextColumn("Nama Obat", width="large"),
            "Stok Terkecil": st.column_config.TextColumn("Stok Terkecil", width="medium"),
            "Golongan": st.column_config.TextColumn("Golongan", width="medium"),
            "Kategori": st.column_config.TextColumn("Kategori", width="medium"),
            "Satuan": st.column_config.TextColumn("Satuan", width="medium"),
            "Status Stok": st.column_config.TextColumn("Status Stok", width="medium"),
            "Worksheet": None
        },
        disabled=["No.", "Kode Obat", "Nama produk", "Stok Terkecil", "Golongan", "Kategori", "Satuan", "Status Stok"],
        key="modal_data_editor_opname_final"
    )

    st.markdown("<br>", unsafe_allow_html=True)
    col_space, col_back = st.columns([8, 2])
    with col_back:
        if st.button("⬅️ Kembali & Masukkan Obat", type="primary", use_container_width=True):
            if edited_modal is not None:
                selected_rows = edited_modal[edited_modal["Pilih"] == True]
                if not selected_rows.empty:
                    items_chosen = df_modal[df_modal["Kode Obat"].isin(selected_rows["Kode Obat"])].copy()
                    st.session_state.opname_custom_items = items_chosen
                else:
                    st.session_state.opname_custom_items = pd.DataFrame()
            st.rerun()

@st.dialog("Konfirmasi Proses Stok Opname", width="large")
def dialog_konfirmasi_proses(edited_opname):
    if "opname_berhasil" not in st.session_state:
        st.session_state.opname_berhasil = False

    if not st.session_state.opname_berhasil:
        st.write("Apakah anda yakin untuk melanjutkan proses stok opname?")
        st.markdown("<br>", unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        if c1.button("Batal", use_container_width=True):
            st.rerun()
        if c2.button("Lanjutkan", type="primary", use_container_width=True):
            ph = st.empty()
            bar = ph.progress(0, text="Mengupload dan memvalidasi data...")
            time.sleep(0.5)
            
            total = len(edited_opname)
            for i, (idx, row) in enumerate(edited_opname.iterrows()):
                bar.progress(int(((i+1)/total)*90), text=f"Memproses item {i+1} dari {total}...")
                time.sleep(0.1) 
                
            do_opname_processing(edited_opname)
            
            bar.progress(100, text="Selesai")
            st.session_state.opname_berhasil = True
            st.rerun()
    else:
        st.success("Stok opname berhasil diproses ke Database!")
        if st.button("Tutup", use_container_width=True):
            st.session_state.opname_berhasil = False
            if "opname_custom_items" in st.session_state:
                del st.session_state.opname_custom_items
            st.rerun()

USERS = {
    "iponadmcantik@gmail.com": {"password": "IponAdmCantik!", "role": "Admin", "name": "Ivonne"},
    "karyawan1@gmail.com": {"password": "karyawan1", "role": "Kasir", "name": "Dian"},
    "karyawan2@gmail.com": {"password": "karyawan2", "role": "Kasir", "name": "Julia"}
}

if "logged_in" not in st.session_state:
    if st.query_params.get("logged_in") == "true":
        st.session_state.logged_in = True
        st.session_state.role = st.query_params.get("role")
        st.session_state.username = st.query_params.get("username")
    else:
        st.session_state.logged_in = False

if "role" not in st.session_state: st.session_state.role = st.query_params.get("role", None)
if "username" not in st.session_state: st.session_state.username = st.query_params.get("username", "")

if not st.session_state.logged_in:
    col1, col2, col3 = st.columns([1, 1.2, 1])
    with col2:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown(
            """
            <div style='text-align:center; padding: 25px 20px 10px 20px; background-color: #16213e; border-radius: 12px 12px 0 0; border: 1px solid #0f3460; border-bottom: none;'>
                <img src='https://img.icons8.com/color/96/pharmacy-shop.png' width='72'/>
                <h2 style='color: #e94560; margin-top: 10px; margin-bottom: 5px;'>Apotek Veteran Blitar</h2>
                <p style='color: #a0a0a0; font-size: 14px; margin-bottom: 0;'>Silakan login untuk melanjutkan</p>
            </div>
            """, unsafe_allow_html=True
        )
        with st.form("form_login"):
            role_pilih = st.selectbox("Login sebagai", ["Admin", "Kasir"])
            username   = st.text_input("Username")
            password   = st.text_input("Password", type="password")
            st.markdown("<br>", unsafe_allow_html=True)
            login_btn  = st.form_submit_button("🔐 Login", use_container_width=True)

            if login_btn:
                uname = username.strip()
                if uname in USERS and USERS[uname]["password"] == password and USERS[uname]["role"] == role_pilih:
                    st.session_state.logged_in = True
                    st.session_state.role = role_pilih
                    st.session_state.username = uname
                    st.query_params["logged_in"] = "true"
                    st.query_params["role"] = role_pilih
                    st.query_params["username"] = uname
                    st.session_state.target_menu = "🏠 Dashboard" 
                    st.query_params["menu"] = "🏠 Dashboard"
                    st.rerun()
                else:
                    st.error("❌ Username, password, atau role tidak sesuai.")
    st.stop()

# ── INIT SESSION STATE UMUM ──
if "retur_items" not in st.session_state:
    st.session_state.retur_items = pd.DataFrame(columns=["Nama produk", "Satuan", "Nomor Batch", "Tanggal Kadaluwarsa", "Stok Sisa", "Jumlah Retur", "Harga 1", "Keterangan"])
if "retur_history" not in st.session_state:
    st.session_state.retur_history = load_retur_history()
if "opname_custom_items" not in st.session_state: st.session_state.opname_custom_items = pd.DataFrame()

# ── INIT SESSION STATE SHIFT ──
if "shift_active" not in st.session_state:
    saved_shift = load_active_shift()
    if saved_shift is not None and saved_shift.get("shift_active") == True:
        st.session_state.shift_active = True
        st.session_state.active_shift_context = {
            "saldo_awal": saved_shift.get("saldo_awal", 0.0),
            "accumulated_sales_expected": saved_shift.get("accumulated_sales_expected", 0.0),
            "start_time": saved_shift.get("start_time"),
            "user_name": saved_shift.get("user_name", ""),
            "joined_users": saved_shift.get("joined_users", []),
            "shift_name": saved_shift.get("shift_name", get_auto_shift_name())
        }
    else:
        st.session_state.shift_active = False
        st.session_state.active_shift_context = {"saldo_awal": 0.0, "accumulated_sales_expected": 0.0, "start_time": None, "user_name": "", "joined_users": [], "shift_name": get_auto_shift_name()}

if "step_tutup_shift" not in st.session_state: st.session_state.step_tutup_shift = 1
if "input_saldo_kasir" not in st.session_state: st.session_state.input_saldo_kasir = 0.0

# ── Sidebar navigasi ──────────────────────────────────────────────────────────
st.sidebar.image("https://img.icons8.com/color/96/pharmacy-shop.png", width=80)
st.sidebar.title("💊 Apotek Veteran Blitar")
st.sidebar.markdown("---")

_role = st.session_state.get("role", "Unknown")
_username = st.session_state.get("username", "")
_name = USERS[_username]["name"] if _username in USERS else "Pengguna"

st.sidebar.markdown(f"👤 **{_name}** — *{_role}*")
st.sidebar.markdown("---")

if _role == "Admin": _menu_options = ["🏠 Dashboard", "📋 Kelola Stok", "🖨️ Rekap Data", "📦 Retur & Entry Pembelian", "🛒 Kasir Utama", "🕒 Sesi Shift"]
else: _menu_options = ["🏠 Dashboard", "📋 Kelola Stok", "🛒 Kasir Utama", "🕒 Sesi Shift"]

if "target_menu" in st.session_state:
    target = st.session_state.target_menu
    del st.session_state.target_menu
    if target in _menu_options: 
        st.session_state.main_menu = target
        st.query_params["menu"] = target

if "main_menu" not in st.session_state:
    saved_menu = st.query_params.get("menu")
    if saved_menu in _menu_options: st.session_state.main_menu = saved_menu
    else: st.session_state.main_menu = _menu_options[0]

def _on_menu_change():
    st.query_params["menu"] = st.session_state.main_menu

menu = st.sidebar.radio("Pilih Fitur", _menu_options, key="main_menu", on_change=_on_menu_change)

if st.sidebar.button("🚪 Logout", use_container_width=True):
    st.session_state.logged_in = False
    st.session_state.role = None
    st.session_state.username = ""
    st.query_params.clear() 
    st.session_state.shift_active = False
    st.session_state.active_shift_context = {"saldo_awal": 0.0, "accumulated_sales_expected": 0.0, "start_time": None, "user_name": "", "joined_users": [], "shift_name": get_auto_shift_name()}
    st.session_state.step_tutup_shift = 1
    st.session_state.input_saldo_kasir = 0.0
    if "main_menu" in st.session_state: del st.session_state["main_menu"]
    st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
if menu == "🏠 Dashboard":
    st.title("💊 Dashboard Apotek Veteran Blitar")
    st.markdown("Selamat datang! Pilih fitur di sidebar untuk mulai mengelola stok obat.")
    st.markdown("---")

    if "inventory_data_cache" not in st.session_state or not st.session_state.inventory_data_cache:
        st.session_state.inventory_data_cache = load_inventory_workbook()

    all_items_df = build_inventory_print_dataframe()
    
    if all_items_df is None or all_items_df.empty:
        st.info("Dataset belum tersedia. Silakan upload dataset di menu **📋 Kelola Stok**.")
    else:
        col_btn_export, _ = st.columns([3, 7])
        with col_btn_export:
            template_buf = io.BytesIO()
            with pd.ExcelWriter(template_buf, engine="openpyxl") as writer:
                export_df = all_items_df[["Nama produk", "Satuan", "Nomor Batch", "Tanggal Kadaluwarsa", "Stok Sisa"]].copy()
                export_df.columns = ["Nama Obat", "Satuan", "No. Batch", "Tanggal Expired", "Stok Sistem"]
                export_df["Stok Nyata Terkecil"] = 0
                export_df["Stok Expired Terkecil"] = 0
                export_df["Satuan Terkecil"] = export_df["Satuan"]
                export_df.to_excel(writer, index=False, sheet_name="Template Stok Opname")
            st.download_button(
                label="📥 Export Template Stok Opname (Excel)", 
                data=template_buf.getvalue(), 
                file_name=f"Template_Stok_Opname_{get_wib_time().date()}.xlsx", 
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                use_container_width=True, 
                type="secondary"
            )

        st.markdown("<br>", unsafe_allow_html=True)
        
        all_items_df["Nama produk"] = all_items_df["Nama produk"].astype(str).str.strip()
        all_items_df = all_items_df[(all_items_df["Nama produk"] != "") & (all_items_df["Nama produk"].str.lower() != "nan") & (all_items_df["Nama produk"].notna())]
        all_items_df["Stok Sisa"] = pd.to_numeric(all_items_df["Stok Sisa"], errors="coerce").fillna(0)
        all_items_df["Harga 1"] = pd.to_numeric(all_items_df["Harga 1"], errors="coerce").fillna(0)
        
        total_jenis = all_items_df["Nama produk"].nunique()
        total_stok = all_items_df["Stok Sisa"].sum()
        
        tgl_batas = pd.Timestamp(get_wib_time().date()) + pd.Timedelta(days=30)
        exp_soon_df = all_items_df[(pd.to_datetime(all_items_df["Tanggal Kadaluwarsa"], errors='coerce') <= tgl_batas) & (all_items_df["Stok Sisa"] > 0)]
        total_exp_soon = exp_soon_df["Nama produk"].nunique()
        
        nilai_stok = (all_items_df["Stok Sisa"] * all_items_df["Harga 1"]).sum()
        
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total Jenis Obat", total_jenis)
        col2.metric("Total Stok Tersedia", f"{int(total_stok):,}".replace(",", "."))
        col3.metric("⚠️ Hampir Kadaluarsa (≤30 hari)", total_exp_soon)
        col4.metric("💰 Estimasi Nilai Stok (Rp)", format_rupiah(nilai_stok))

        st.markdown("---")
        col_low, col_exp = st.columns(2)
        with col_low:
            st.markdown("#### 📉 Stok Menipis (≤ 20)")
            cari_low = st.text_input("🔍 Cari (Nama, Batch, PBF, dll)", key="cari_low", placeholder="Cari obat stok menipis...")
            stok_df = all_items_df.copy()
            if cari_low.strip():
                mask_low = stok_df.astype(str).apply(lambda col: col.str.contains(cari_low.strip(), case=False, na=False)).any(axis=1)
                stok_df = stok_df[mask_low]
            stok_df["Worksheet"] = stok_df["Worksheet"].fillna("-")
            stok_summary = stok_df.groupby(["Worksheet", "Nama produk"])["Stok Sisa"].sum().reset_index()
            stok_menipis = stok_summary[stok_summary["Stok Sisa"] <= 20].sort_values("Stok Sisa")
            if stok_menipis.empty: st.success("Tidak ada obat dengan stok menipis atau yang cocok dengan pencarian.")
            else: st.dataframe(stok_menipis.rename(columns={"Nama produk": "Nama Obat", "Stok Sisa": "Total Stok"}), use_container_width=True, hide_index=True)
                
        with col_exp:
            st.markdown("#### ⏰ Segera Kadaluarsa (≤30 hari)")
            cari_exp = st.text_input("🔍 Cari (Nama, Batch, PBF, dll)", key="cari_exp", placeholder="Cari obat segera kadaluarsa...")
            exp_df = exp_soon_df.copy()
            if cari_exp.strip():
                mask_exp = exp_df.astype(str).apply(lambda col: col.str.contains(cari_exp.strip(), case=False, na=False)).any(axis=1)
                exp_df = exp_df[mask_exp]
            if exp_df.empty: st.success("Tidak ada obat yang mendekati tanggal kadaluarsa atau yang cocok dengan pencarian.")
            else:
                exp_show = exp_df[["Nama produk", "Worksheet", "Nomor Batch", "Tanggal Kadaluwarsa", "Stok Sisa"]].copy()
                exp_show["Tanggal Kadaluwarsa"] = exp_show["Tanggal Kadaluwarsa"].apply(lambda x: x.strftime("%d-%m-%Y") if pd.notna(x) else "")
                st.dataframe(exp_show.rename(columns={"Nama produk": "Nama Obat", "Tanggal Kadaluwarsa": "Tgl Expired", "Nomor Batch": "Batch"}), use_container_width=True, hide_index=True)

# ══════════════════════════════════════════════════════════════════════════════
# KELOLA STOK (DATABASE & BACKUP) - PERBAIKAN PENDETEKSIAN HEADER
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "📋 Kelola Stok":
    st.title("📋 Kelola Stok")
    if not HAS_SQLALCHEMY:
        st.info("ℹ️ Mode Database Lokal Aktif. Tambahkan `SQLAlchemy` dan `psycopg2-binary` ke `requirements.txt` jika ingin mengaktifkan Cloud Database (Supabase).")
    else:
        st.caption("Kelola dan sesuaikan data stok obat secara langsung atau melalui Stok Opname.")

    if "inventory_data_cache" not in st.session_state: st.session_state.inventory_data_cache = {}

    if st.session_state.role == "Admin":
        col_up, col_dl = st.columns([7, 3])
        with col_dl:
            st.markdown("<div style='margin-top: 25px;'></div>", unsafe_allow_html=True)
            zip_data = get_backup_zip()
            st.download_button(
                label="📥 Download Backup Database (.ZIP)",
                data=zip_data,
                file_name=f"Backup_DB_Apotek_Veteran_{get_wib_time().strftime('%Y%m%d_%H%M%S')}.zip",
                mime="application/zip",
                use_container_width=True,
                type="primary"
            )
            
        with col_up:
            uploaded_inventory = st.file_uploader("Upload file Excel/CSV langsung dari perangkat Anda (untuk menimpa Database Master):", type=["xlsx", "xlsm", "csv"], key="upload_inventory_source")
            if uploaded_inventory is not None:
                file_id = f"{uploaded_inventory.name}_{uploaded_inventory.size}"
                if st.session_state.get("last_uploaded_file") != file_id:
                    try:
                        filename_lower = uploaded_inventory.name.lower()
                        # Tangani CSV
                        if filename_lower.endswith(".csv"):
                            uploaded_inventory.seek(0)
                            try:
                                df = pd.read_csv(uploaded_inventory, sep=None, engine='python')
                            except Exception:
                                uploaded_inventory.seek(0)
                                df = pd.read_csv(uploaded_inventory, sep=';')
                                
                            if "Worksheet" not in df.columns: df["Worksheet"] = "TAB"
                            
                            df = prepare_sheet_for_editor(df)
                            for col in df.columns:
                                df[col] = df[col].apply(lambda v: str(v) if isinstance(v, (list, tuple, dict, set)) else v)
                            db_write_table(df, "inventory_master")
                        else:
                            # [PERBAIKAN] Menggunakan openpyxl load_workbook untuk memindai baris header dengan cerdas!
                            wb = load_workbook(io.BytesIO(uploaded_inventory.getvalue()), data_only=True)
                            frames = []
                            for sheet in wb.sheetnames:
                                ws = wb[sheet]
                                df_sheet = load_inventory_sheet_dataframe(ws)
                                
                                # Lewati sheet yang kosong atau tidak valid
                                if df_sheet.empty or df_sheet["Nama produk"].isnull().all():
                                    continue
                                    
                                df_sheet = prepare_sheet_for_editor(df_sheet)
                                df_sheet["Worksheet"] = sheet
                                frames.append(df_sheet)
                                
                            if frames:
                                combined = pd.concat(frames, ignore_index=True)
                                combined = combined.dropna(subset=["Nama produk"], how="all")
                                for col in combined.columns:
                                    combined[col] = combined[col].apply(lambda v: str(v) if isinstance(v, (list, tuple, dict, set)) else v)
                                db_write_table(combined, "inventory_master")
                        
                        st.session_state.inventory_data_cache = load_inventory_workbook()
                        st.session_state["last_uploaded_file"] = file_id
                        st.toast("✅ Data berhasil dimasukkan permanen ke Database!")
                        time.sleep(1)
                        st.rerun()
                    except Exception as e:
                        st.error(f"Gagal memproses file ke Database: {e}")

    workbook_data = st.session_state.inventory_data_cache
    if not workbook_data:
        workbook_data = load_inventory_workbook()
        st.session_state.inventory_data_cache = workbook_data

    if not workbook_data: st.info("Database kosong. Silakan upload file Excel/CSV pada kotak di atas.")
    AVAILABLE_SHEETS = get_available_sheets()

    tab_edit, tab_opname = st.tabs(["✏️ Edit Stok", "📦 Stok Opname"])

    with tab_edit:
        if workbook_data:
            sheet_name = st.selectbox("Pilih Worksheet", AVAILABLE_SHEETS, index=0, key="inventory_selected_sheet")
            if sheet_name not in workbook_data:
                sheet_df = pd.DataFrame(columns=INVENTORY_COLUMNS)
                sheet_df = prepare_sheet_for_editor(sheet_df)
            else:
                sheet_df = prepare_sheet_for_editor(workbook_data[sheet_name].copy())

            if st.session_state.role == "Admin": st.info("Setiap kolom dalam tabel dapat diedit langsung dengan ikon ✏️.")
            else: st.info("Cari data stok di bawah ini. Anda hanya dapat melihat data (Read-Only).")
                
            search_inv = st.text_input("🔍 Pencarian Baris (Nama, Batch, Faktur, PBF, dll di Worksheet ini)", placeholder="Ketik kata kunci...")
            if search_inv.strip():
                mask = sheet_df.astype(str).apply(lambda col: col.str.contains(search_inv.strip(), case=False, na=False)).any(axis=1)
                display_df = sheet_df[mask].copy()
            else: display_df = sheet_df.copy()

            if st.session_state.role == "Admin":
                editor_cols = [c for c in INVENTORY_COLUMNS if c != "Worksheet"]
                edited_display_df = st.data_editor(
                    display_df, use_container_width=True, num_rows="dynamic", hide_index=True, column_order=editor_cols,
                    column_config={
                        "Nama produk": st.column_config.TextColumn("✏️ Nama Produk", width="large"),
                        "Tanggal": st.column_config.DateColumn("✏️ Tanggal", format="YYYY-MM-DD", width="medium"),
                        "Tanggal Kadaluwarsa": st.column_config.DateColumn("✏️ Tanggal Kadaluarsa", format="YYYY-MM-DD", width="medium"),
                        "Stok Masuk": st.column_config.NumberColumn("✏️ Stok Masuk", min_value=0, step=1, width="small"),
                        "Stok Keluar": st.column_config.NumberColumn("✏️ Stok Keluar", min_value=0, step=1, width="small"),
                        "Stok Sisa": st.column_config.NumberColumn("✏️ Stok Sisa", min_value=0, step=1, width="small"),
                    }, key="editor_inventory_grid"
                )

                if st.button("✅ Submit Data Terbaru", type="primary"):
                    current_ws_df = prepare_sheet_for_editor(workbook_data[sheet_name].copy())
                    existing_idx = edited_display_df.index.intersection(current_ws_df.index)
                    current_ws_df.loc[existing_idx, edited_display_df.columns] = edited_display_df.loc[existing_idx]
                    
                    new_rows = edited_display_df[~edited_display_df.index.isin(current_ws_df.index)]
                    if not new_rows.empty: current_ws_df = pd.concat([current_ws_df, new_rows])
                        
                    deleted_rows = display_df.index.difference(edited_display_df.index)
                    if not deleted_rows.empty: current_ws_df = current_ws_df.drop(deleted_rows)
                        
                    current_ws_df = current_ws_df.reset_index(drop=True)
                    workbook_data[sheet_name] = normalize_inventory_df(current_ws_df)
                    if save_inventory_workbook(workbook_data):
                        st.session_state.inventory_data_cache = workbook_data
                        st.toast(f"✅ Perubahan pada worksheet {sheet_name} berhasil disimpan ke Database.")
                        time.sleep(1)
                        st.rerun()
            else:
                editor_cols = [c for c in INVENTORY_COLUMNS if c != "Worksheet"]
                st.dataframe(display_df, use_container_width=True, hide_index=True, column_order=editor_cols,
                    column_config={"Tanggal": st.column_config.DateColumn("Tanggal", format="YYYY-MM-DD"), "Tanggal Kadaluwarsa": st.column_config.DateColumn("Tanggal Kadaluarsa", format="YYYY-MM-DD")},
                    key="viewer_inventory_grid")

    with tab_opname:
        if workbook_data:
            st.markdown("<h3 style='text-align: center; color: #e94560; margin-bottom: 20px;'>Stok Opname Obat</h3>", unsafe_allow_html=True)
            if st.session_state.role != "Admin": st.info("Fitur Stok Opname hanya dapat diakses dan diproses oleh Admin. Tampilan di bawah ini bersifat Read-Only.")

            c_file1, c_file2, c_file3 = st.columns([5, 2, 3])
            with c_file1: st.file_uploader("Upload File Opname", type=["xlsx", "csv"], key="import_opname_file", label_visibility="collapsed")
            with c_file2: 
                st.markdown("<div style='margin-top: 2px;'></div>", unsafe_allow_html=True); st.button("Cari File", key="btn_cari_file_opname", use_container_width=True)
            with c_file3: 
                st.markdown("<div style='margin-top: 2px;'></div>", unsafe_allow_html=True); st.button("Import Stok Opname", key="btn_import_opname", use_container_width=True)

            st.markdown("<div style='margin-bottom: 10px;'></div>", unsafe_allow_html=True)

            c_gudang, c_proses, c_reset, _ = st.columns([3, 2, 2, 5])
            with c_gudang: opname_gudang = st.selectbox("Pilih Gudang", AVAILABLE_SHEETS, key="opname_gudang_select", label_visibility="collapsed")
            with c_proses: btn_proses_opname = st.button("✅ Proses", use_container_width=True, type="primary", key="btn_proses_opname_action")
            with c_reset: btn_reset_opname = st.button("🔄 Reset", use_container_width=True, key="btn_reset_opname_action")

            st.markdown("<div style='margin-bottom: 10px;'></div>", unsafe_allow_html=True)

            c_search, c_btn_cari = st.columns([9, 1])
            with c_search: search_opname = st.text_input("Pencarian obat", placeholder="Ketik kata kunci untuk mencari...", label_visibility="collapsed", key="search_opname_input")
            with c_btn_cari: btn_cari_obat = st.button("Cari", type="primary", use_container_width=True, key="btn_cari_obat_opname_trigger")

            if btn_cari_obat:
                df_all = build_inventory_print_dataframe()
                if df_all is not None and not df_all.empty: modal_pilih_obat(df_all, search_opname)

            df_ws_opname = pd.DataFrame()
            if opname_gudang in workbook_data:
                df_ws_opname = workbook_data[opname_gudang].copy()
                df_ws_opname = prepare_sheet_for_editor(df_ws_opname)
                
            df_opname = pd.DataFrame()
            if not df_ws_opname.empty:
                df_opname["Worksheet"] = opname_gudang
                df_opname["No."] = range(1, len(df_ws_opname) + 1)
                df_opname["Kode Obat"] = ["OBT" + str(1000 + i) for i in range(len(df_ws_opname))]
                df_opname["Nama Obat"] = df_ws_opname["Nama produk"]
                df_opname["Satuan"] = df_ws_opname["Satuan"]
                df_opname["No. Batch"] = df_ws_opname["Nomor Batch"]
                df_opname["Tanggal Expired"] = df_ws_opname["Tanggal Kadaluwarsa"].apply(lambda x: x.strftime("%d %b %Y") if pd.notna(x) else "-")
                df_opname["Stok Satuan Terkecil"] = pd.to_numeric(df_ws_opname["Stok Sisa"], errors="coerce").fillna(0)
                df_opname["Stok Nyata Terkecil"] = 0.00
                df_opname["Stok Expired Terkecil"] = 0.00
                df_opname["Satuan Terkecil"] = df_ws_opname["Satuan"]

            if search_opname.strip() and not btn_cari_obat:
                mask_opname = df_opname.astype(str).apply(lambda col: col.str.contains(search_opname.strip(), case=False, na=False)).any(axis=1)
                df_opname = df_opname[mask_opname]

            if "opname_custom_items" in st.session_state and not st.session_state.opname_custom_items.empty:
                chosen = st.session_state.opname_custom_items
                df_opname = pd.DataFrame()
                df_opname["Worksheet"] = chosen["Worksheet"].values
                df_opname["No."] = range(1, len(chosen) + 1)
                df_opname["Kode Obat"] = chosen["Kode Obat"].values
                df_opname["Nama Obat"] = chosen["Nama produk"].values
                df_opname["Satuan"] = chosen["Satuan"].values
                df_opname["No. Batch"] = chosen["Nomor Batch"].values
                df_opname["Tanggal Expired"] = chosen["Tanggal Kadaluwarsa"].apply(lambda x: x.strftime("%d %b %Y") if pd.notna(x) else "-").values
                df_opname["Stok Satuan Terkecil"] = pd.to_numeric(chosen["Stok Sisa"], errors="coerce").fillna(0).values
                df_opname["Stok Nyata Terkecil"] = 0.00
                df_opname["Stok Expired Terkecil"] = 0.00
                df_opname["Satuan Terkecil"] = chosen["Satuan"].values

            st.caption(f"Menampilkan 1-{len(df_opname)} dari {len(df_opname)} data")

            if st.session_state.role == "Admin":
                edited_opname = st.data_editor(
                    df_opname, use_container_width=True, hide_index=True,
                    disabled=["No.", "Kode Obat", "Nama Obat", "Satuan", "No. Batch", "Tanggal Expired", "Stok Satuan Terkecil", "Satuan Terkecil"],
                    column_config={"Worksheet": None, "Stok Nyata Terkecil": st.column_config.NumberColumn("Stok Nyata Terkecil", min_value=0.0, step=1.0, format="%.2f"), "Stok Expired Terkecil": st.column_config.NumberColumn("Stok Expired Terkecil", min_value=0.0, step=1.0, format="%.2f")},
                    key="opname_editor_grid_final"
                )
                
                if btn_proses_opname:
                    if edited_opname is not None and not edited_opname.empty:
                        dialog_konfirmasi_proses(edited_opname)

                if btn_reset_opname:
                    if "opname_custom_items" in st.session_state: del st.session_state.opname_custom_items
                    st.rerun()
            else:
                if "Worksheet" in df_opname.columns: st.dataframe(df_opname.drop(columns=["Worksheet"]), use_container_width=True, hide_index=True)
                else: st.dataframe(df_opname, use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# FITUR REKAP DATA (DATABASE + AUTO CLEANER)
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "🖨️ Rekap Data":
    st.title("🖨️ Rekap Data")

    df_history = load_data()

    if df_history is not None and not df_history.empty:
        mask_legacy = df_history["Keterangan"].astype(str).str.contains("Kasir Pembelian Obat", case=False, na=False)
        if mask_legacy.any():
            df_history = df_history[~mask_legacy]
            save_data(df_history)

    if df_history is None or df_history.empty:
        st.warning("Belum ada data riwayat transaksi (Penjualan, Opname, atau Entry Pembelian). Silakan lakukan transaksi terlebih dahulu.")
        st.stop()

    df_history["Tanggal"] = pd.to_datetime(df_history["Tanggal"], errors="coerce")
    if "Tanggal Kadaluarsa" in df_history.columns:
        df_history["Tanggal Kadaluarsa"] = pd.to_datetime(df_history["Tanggal Kadaluarsa"], errors="coerce")

    all_items_df = build_inventory_print_dataframe()
    obat_excel_list = []
    if all_items_df is not None and not all_items_df.empty:
        obat_excel_list = [str(x).strip() for x in all_items_df["Nama produk"].dropna().unique() if str(x).strip()]
    obat_history_list = [str(x).strip() for x in df_history["Nama Obat"].dropna().unique() if str(x).strip()]
    
    pilihan_obat = ["SEMUA OBAT"] + sorted(list(set(obat_excel_list + obat_history_list)))
    satuan_list = sorted([str(x).strip() for x in df_history["Satuan"].dropna().unique() if str(x).strip()])
    pilihan_satuan = ["SEMUA SATUAN"] + satuan_list

    with st.container():
        st.markdown("<div class='filter-label'>Pilihan Periode</div>", unsafe_allow_html=True)
        c1, c2, c3, c4 = st.columns([2, 2, 2, 4])
        with c1:
            periode_opt = st.selectbox("Pilih Periode", ["Semua Waktu", "Berdasarkan Tanggal", "Berdasarkan Bulan", "Berdasarkan Tahun"], label_visibility="collapsed")
        
        tgl_awal, tgl_akhir = None, None
        periode_text = "SEMUA WAKTU"

        if periode_opt == "Berdasarkan Tanggal":
            with c2:
                tgl_range = st.date_input("Rentang Tanggal", value=(get_wib_time().date(), get_wib_time().date()), label_visibility="collapsed")
            if len(tgl_range) == 2:
                tgl_awal, tgl_akhir = tgl_range
                periode_text = f"{tgl_awal.strftime('%d %b %Y')} - {tgl_akhir.strftime('%d %b %Y')}"
            else:
                st.warning("Pilih rentang tanggal yang lengkap.")
                st.stop()
        elif periode_opt == "Berdasarkan Bulan":
            with c2: bln = st.selectbox("Bulan", range(1, 13), index=get_wib_time().month-1, label_visibility="collapsed")
            with c3: thn = st.number_input("Tahun", min_value=2000, max_value=2100, value=get_wib_time().year, label_visibility="collapsed")
            tgl_awal = pd.Timestamp(year=thn, month=bln, day=1).date()
            tgl_akhir = (pd.Timestamp(year=thn, month=bln, day=1) + pd.offsets.MonthEnd(0)).date()
            periode_text = f"BULAN {bln} TAHUN {thn}"
        elif periode_opt == "Berdasarkan Tahun":
            with c2: thn_only = st.number_input("Tahun", min_value=2000, max_value=2100, value=get_wib_time().year, label_visibility="collapsed")
            tgl_awal = pd.Timestamp(year=thn_only, month=1, day=1).date()
            tgl_akhir = pd.Timestamp(year=thn_only, month=12, day=31).date()
            periode_text = f"TAHUN {thn_only}"

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("<div class='filter-label'>Filter Spesifik Kolom</div>", unsafe_allow_html=True)
        c_f1, c_f2, c_f3 = st.columns(3)
        with c_f1: filter_satuan = st.selectbox("Satuan Obat", options=pilihan_satuan)
        with c_f2: filter_nama = st.selectbox("Pilih Obat", options=pilihan_obat)
        with c_f3: filter_ket = st.text_input("Keterangan", placeholder="Cari Keterangan/Faktur...")

    df_print = df_history.copy()
    
    if tgl_awal and tgl_akhir:
        df_print = df_print[(df_print["Tanggal"].dt.date >= tgl_awal) & (df_print["Tanggal"].dt.date <= tgl_akhir)]

    if filter_satuan != "SEMUA SATUAN":
        df_print = df_print[df_print["Satuan"].astype(str).str.strip() == filter_satuan]
        
    if filter_nama != "SEMUA OBAT":
        df_print = df_print[df_print["Nama Obat"].astype(str).str.strip() == filter_nama]

    if filter_ket.strip():
        mask_ket = df_print["Keterangan"].astype(str).str.contains(filter_ket.strip(), case=False, na=False)
        mask_faktur = df_print["Nomor Faktur"].astype(str).str.contains(filter_ket.strip(), case=False, na=False)
        df_print = df_print[mask_ket | mask_faktur]

    preview_df = df_print.copy()
    
    preview_df = preview_df[[
        "Tanggal", "Nama Obat", "Keterangan", "Stok Masuk", "Stok Keluar", 
        "Stok Akhir", "Satuan", "Nomor Batch", "Tanggal Kadaluarsa", "Petugas"
    ]]
    
    preview_df["Nomor Batch"] = preview_df["Nomor Batch"].fillna("-")
    preview_df["Petugas"] = preview_df["Petugas"].fillna("-")
    
    preview_df.rename(columns={
        "Stok Masuk": "Masuk",
        "Stok Keluar": "Keluar",
        "Stok Akhir": "Sisa",
        "Tanggal Kadaluarsa": "Tanggal Expired"
    }, inplace=True)
    
    if "Tanggal" in preview_df.columns: 
        preview_df["Tanggal"] = preview_df["Tanggal"].apply(lambda x: x.strftime("%d %b %Y %H:%M:%S") if pd.notna(x) else "")
    if "Tanggal Expired" in preview_df.columns: 
        preview_df["Tanggal Expired"] = preview_df["Tanggal Expired"].apply(lambda x: x.strftime("%d %b %Y") if pd.notna(x) else "")

    st.markdown("<br>", unsafe_allow_html=True)
    c_action1, c_action2, c_action3 = st.columns([2, 2, 6])
    with c_action1:
        st.button("🔍 Terapkan Filter", type="primary", use_container_width=True)
        
    html_rows = ""
    for i, row in enumerate(preview_df.values):
        html_rows += f"<tr><td style='text-align:center;'>{i+1}</td>" + "".join(f"<td>{v}</td>" for v in row) + "</tr>"
    html_headers = "<th>No.</th>" + "".join(f"<th>{c}</th>" for c in preview_df.columns)

    html_printable_laporan = f"""
    <!DOCTYPE html>
    <html><head><meta charset='utf-8'><title>Laporan Kartu Stok Obat - Apotek Veteran Blitar</title>
    <style>
        @page {{ size: A4 landscape; margin: 15mm; }}
        body {{ font-family: 'Times New Roman', Times, serif; font-size: 12pt; margin: 0; padding: 0; background: #fff; color: #000; }}
        .header {{ text-align: center; border-bottom: 2pt solid #000; padding-bottom: 10px; margin-bottom: 20px; }}
        .header h3 {{ margin: 0 0 5px 0; font-size: 14pt; font-weight: bold; text-transform: uppercase; }}
        .header p {{ margin: 0; font-size: 12pt; }}
        .title {{ text-align: center; font-size: 14pt; font-weight: bold; margin-bottom: 5px; text-decoration: underline; }}
        .subtitle {{ text-align: center; font-size: 12pt; font-weight: bold; margin-bottom: 20px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 10px; color: #000; }}
        th, td {{ border: 1pt solid #000; padding: 6px 8px; text-align: left; font-size: 12pt; }}
        th {{ text-align: center; font-weight: bold; background-color: #fff; }}
        .info-table {{ width: 100%; border: none; margin-bottom: 15px; font-size: 12pt; }}
        .info-table td {{ border: none; padding: 2px 5px; font-weight: bold; }}
    </style></head>
    <body>
        <div class='header'>
            <h3>APOTEK VETERAN SEHAT BLITAR</h3>
            <p>No. Surat Izin Apotek : 30032300193080001</p>
            <p>Jl. Veteran No 64B Kota Blitar, Kota Blitar</p>
            <p>Telp. 081331808585, Email : veteransehat01@gmail.com</p>
        </div>
        <div class='title'>LAPORAN KARTU STOK OBAT</div>
        <div class='subtitle'>PERIODE {periode_text}</div>
        
        <table class='info-table'>
            <tr>
                <td style='width: 15%;'>Satuan Obat</td><td style='width: 35%;'>: {filter_satuan}</td>
                <td style='width: 15%;'></td><td style='width: 35%;'></td>
            </tr>
            <tr>
                <td>Nama Obat</td><td>: {filter_nama}</td>
                <td></td><td></td>
            </tr>
        </table>

        <table>
            <thead><tr>{html_headers}</tr></thead>
            <tbody>{html_rows}</tbody>
        </table>
    </body></html>
    """

    b64_html_laporan = base64.b64encode(html_printable_laporan.encode("utf-8")).decode("utf-8")
    
    with c_action2:
        custom_print_laporan = f"""
        <!DOCTYPE html><html><head><style>
            body {{ margin: 0; padding: 0; font-family: 'Segoe UI', sans-serif; background: transparent; }}
            .btn {{ display: flex; align-items: center; justify-content: center; width: 100%; height: 38px; background-color: #28a745; color: white; border: none; border-radius: 8px; font-size: 14px; font-weight: 600; cursor: pointer; transition: 0.3s; }}
            .btn:hover {{ background-color: #218838; }}
        </style></head><body>
            <button class="btn" onclick="printReport()">🖨️ Cetak Laporan (Print)</button>
            <script>
            function printReport() {{
                const b64 = "{b64_html_laporan}";
                const binStr = atob(b64);
                const bytes = new Uint8Array(binStr.length);
                for (let i = 0; i < binStr.length; i++) {{ bytes[i] = binStr.charCodeAt(i); }}
                const htmlContent = new TextDecoder('utf-8').decode(bytes);
                const printWin = window.open('', '_blank');
                printWin.document.open(); 
                printWin.document.write(htmlContent); 
                printWin.document.close();
                setTimeout(function() {{ printWin.focus(); printWin.print(); }}, 500);
            }}
            </script>
        </body></html>
        """
        components.html(custom_print_laporan, height=45)

    st.markdown("---")
    st.caption(f"Menampilkan {len(preview_df)} data dari total {len(df_history)} riwayat transaksi.")
    st.dataframe(preview_df, use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# KASIR UTAMA
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "🛒 Kasir Utama":
    
    st.title("🛒 Kasir Utama")
    
    if not st.session_state.shift_active:
        st.error("⚠️ Anda belum membuka shift! Buka shift terlebih dahulu agar transaksi kasir dapat direkap.")
        if st.button("🕒 Menuju Halaman Buka Shift", type="primary"):
            st.session_state.target_menu = "🕒 Sesi Shift"
            st.rerun()
        st.stop()

    start_time_str = st.session_state.active_shift_context.get("start_time")
    if start_time_str:
        start_date = start_time_str.split(" ")[0]
        if start_date != get_wib_time().date().strftime("%Y-%m-%d"):
            st.error("⚠️ Terdapat shift dari hari sebelumnya yang belum ditutup. Tutup shift hari sebelumnya di menu 'Sesi Shift' terlebih dahulu sebelum memulai transaksi hari ini atau pulang!")
            if st.button("Pindah ke Menu Sesi Shift", type="primary"):
                st.session_state.target_menu = "🕒 Sesi Shift"
                st.rerun()
            st.stop()

    active_ctx = st.session_state.active_shift_context
    valid_users = [active_ctx.get("user_name", "")] + active_ctx.get("joined_users", [])
    valid_users = [u for u in valid_users if u]
    
    current_kasir_name = USERS.get(st.session_state.username, {}).get("name", "Unknown")
    
    if current_kasir_name not in valid_users and st.session_state.role != "Admin": 
        st.warning(f"⚠️ Saudara/i {current_kasir_name}, Anda belum tergabung dalam daftar Shift aktif ini. Masuk ke menu 'Sesi Shift' lalu klik 'Gabung Shift Ini'.")
        if st.button("Pindah ke Menu Sesi Shift", type="primary"):
            st.session_state.target_menu = "🕒 Sesi Shift"
            st.rerun()
        st.stop()

    if "inventory_data_cache" not in st.session_state or not st.session_state.inventory_data_cache:
        st.warning("Dataset Excel belum tersedia. Silakan upload terlebih dahulu di menu **📋 Kelola Stok**.")
        st.stop()
        
    all_items_df = build_inventory_print_dataframe()
    if all_items_df is None or all_items_df.empty:
        st.warning("Data stok kosong.")
        st.stop()

    if "cart" not in st.session_state: st.session_state.cart = []
    if "checkout_mode" not in st.session_state: st.session_state.checkout_mode = False
    if "bayar_tunai" not in st.session_state: st.session_state.bayar_tunai = 0
    if "nota_confirmed" not in st.session_state: st.session_state.nota_confirmed = False

    col_input, col_nota = st.columns([1, 1])

    with col_input:
        st.subheader("🛒 Input Penjualan")
        
        pilihan_kasir = valid_users if valid_users else [current_kasir_name]
        if current_kasir_name in pilihan_kasir:
            default_kasir_idx = pilihan_kasir.index(current_kasir_name)
        else:
            default_kasir_idx = 0
            
        kasir_aktif = st.selectbox("👩‍💻 Pilih Kasir yang Bertugas:", pilihan_kasir, index=default_kasir_idx)
        kasir_nama_nota = kasir_aktif
        
        st.caption("Penjualan memotong stok secara real-time dari Database.")

        available_items = all_items_df[all_items_df["Stok Sisa"].fillna(0) > 0].copy()
        if available_items.empty:
            st.info("Tidak ada obat dengan stok tersedia (>0).")
        else:
            available_items["Label"] = available_items.apply(
                lambda x: f"{str(x['Nama produk']).strip()} | {str(x['Satuan']).strip() if pd.notna(x['Satuan']) and str(x['Satuan']).strip() != '' else str(x['Worksheet']).strip()} | Stok: {int(x['Stok Sisa'])}",
                axis=1
            )

            if not st.session_state.checkout_mode:
                selected_label = st.selectbox("Pilih Obat (Bisa diketik untuk mencari)", available_items["Label"].unique().tolist(), key="kasir_pilih_obat")
                selected_row_display = available_items[available_items["Label"] == selected_label].iloc[0]
                satuan_display = str(selected_row_display["Satuan"]).strip() if pd.notna(selected_row_display["Satuan"]) and str(selected_row_display["Satuan"]).strip() != "" else str(selected_row_display["Worksheet"]).strip()

                with st.form("form_kasir"):
                    col_su, col_sh = st.columns(2)
                    with col_su: st.text_input("Satuan Jual", value=satuan_display, disabled=True)
                    with col_sh: skema_harga = st.selectbox("Skema Harga", ["Harga 1", "Harga 2"])

                    jumlah = st.number_input("Jumlah", min_value=1, value=1)
                    add_to_cart = st.form_submit_button("➕ Tambah ke Nota")

                    if add_to_cart:
                        selected_row = available_items[available_items["Label"] == selected_label].iloc[0]
                        nama_obat = selected_row["Nama produk"]
                        ws_target = selected_row["Worksheet"]
                        batch_target = selected_row["Nomor Batch"]
                        satuan_jual = satuan_display
                        
                        harga_per_satuan = float(selected_row["Harga 1"]) if skema_harga == "Harga 1" else float(selected_row["Harga 2"])
                        if pd.isna(harga_per_satuan): harga_per_satuan = 0.0
                        
                        subtotal = harga_per_satuan * jumlah
                        stok_tersedia = float(selected_row["Stok Sisa"])

                        if jumlah > stok_tersedia:
                            st.error(f"❌ Stok tidak cukup! Tersedia {stok_tersedia:.0f}, dibutuhkan {jumlah:.0f}.")
                        else:
                            st.session_state.cart.append({
                                "nama": nama_obat, "worksheet": ws_target, "batch": batch_target,
                                "satuan_jual": satuan_jual, "qty": jumlah, "skema_harga": skema_harga,
                                "harga_per_satuan": harga_per_satuan, "subtotal": subtotal, "tgl_exp": selected_row["Tanggal Kadaluwarsa"]
                            })
                            st.toast(f"✅ {nama_obat} ({jumlah} {satuan_jual}) ditambah ke nota!")

                if st.session_state.cart:
                    st.markdown("---")
                    st.markdown("#### 🛒 Rincian Keranjang")
                    for i, item in enumerate(st.session_state.cart):
                        c1, c2, c3, c4 = st.columns([4, 2, 2, 2])
                        c1.write(f"**{item['nama']}**")
                        c2.write(f"x{item['qty']} {item['satuan_jual']}")
                        c3.write(format_rupiah(item['subtotal']))
                        with c4:
                            col_min, col_del = st.columns(2)
                            with col_min:
                                if st.button("➖", key=f"min_{i}", help="Kurangi 1"):
                                    if st.session_state.cart[i]["qty"] > 1:
                                        st.session_state.cart[i]["qty"] -= 1
                                        st.session_state.cart[i]["subtotal"] = st.session_state.cart[i]["harga_per_satuan"] * st.session_state.cart[i]["qty"]
                                    else: st.session_state.cart.pop(i)
                                    st.rerun()
                            with col_del:
                                if st.button("🗑️", key=f"del_{i}", help="Hapus item"):
                                    st.session_state.cart.pop(i)
                                    st.rerun()
                    st.markdown("")
                    if st.button("✅ Lanjut ke Pembayaran", type="primary", use_container_width=True):
                        st.session_state.checkout_mode = True
                        st.rerun()
            else:
                if not st.session_state.nota_confirmed:
                    st.info(f"🛒 **{len(st.session_state.cart)} item** dalam keranjang. Silakan masukkan nominal bayar.")
                    bayar_input = st.number_input("Nominal Bayar Uang Fisik (Rp)", min_value=0, step=500, value=st.session_state.bayar_tunai)
                    st.session_state.bayar_tunai = bayar_input

                    st.markdown("<br>", unsafe_allow_html=True)
                    col_teliti, col_submit = st.columns(2)
                    with col_teliti:
                        if st.button("🔍 Teliti Kembali Keranjang", type="secondary", use_container_width=True):
                            st.session_state.checkout_mode = False
                            st.session_state.nota_confirmed = False
                            st.rerun()
                    with col_submit:
                        if st.button("✅ Konfirmasi Pembayaran", type="primary", use_container_width=True):
                            if st.session_state.bayar_tunai <= 0: st.error("Nominal bayar harus diisi!")
                            else:
                                workbook_data = st.session_state.inventory_data_cache
                                df_history = load_data()
                                new_history_rows = []
                                total_belanja_confirm = 0.0
                                
                                invoice_no = f"INV-{get_wib_time().strftime('%y%m%d%H%M%S')}"

                                for item in st.session_state.cart:
                                    ws_target = item["worksheet"]
                                    if ws_target in workbook_data:
                                        sheet_df = prepare_sheet_for_editor(workbook_data[ws_target].copy())
                                        mask = ((sheet_df["Nama produk"].fillna("").astype(str) == str(item["nama"])) & (sheet_df["Nomor Batch"].fillna("").astype(str) == str(item["batch"])))
                                        if mask.any():
                                            idx = sheet_df[mask].index[-1]
                                            sisa_lama = float(sheet_df.loc[idx, "Stok Sisa"]) if pd.notna(sheet_df.loc[idx, "Stok Sisa"]) else 0.0
                                            keluar_lama = float(sheet_df.loc[idx, "Stok Keluar"]) if pd.notna(sheet_df.loc[idx, "Stok Keluar"]) else 0.0
                                            sisa_baru = max(sisa_lama - item["qty"], 0)
                                            keluar_baru = keluar_lama + item["qty"]
                                            sheet_df.loc[idx, "Stok Sisa"] = sisa_baru
                                            sheet_df.loc[idx, "Stok Keluar"] = keluar_baru
                                            workbook_data[ws_target] = sheet_df

                                    total_belanja_confirm += item["subtotal"]
                                    
                                    new_history_rows.append({
                                        "Tanggal": get_wib_time().strftime("%Y-%m-%d %H:%M:%S"),
                                        "Nomor Faktur": invoice_no,
                                        "Nama Obat": item["nama"], 
                                        "Kategori": ws_target, 
                                        "Satuan": item["satuan_jual"], 
                                        "Nomor Batch": item["batch"],
                                        "Stok Masuk": 0, "Stok Keluar": item["qty"], "Stok Akhir": sisa_baru if 'sisa_baru' in locals() else 0,
                                        "Harga Satuan (Rp)": item["harga_per_satuan"], "Total Nilai (Rp)": item["subtotal"],
                                        "Tanggal Kadaluarsa": pd.Timestamp(item["tgl_exp"]) if pd.notna(item["tgl_exp"]) else pd.NaT,
                                        "Keterangan": f"Penjualan dengan No Faktur {invoice_no}",
                                        "Petugas": kasir_nama_nota
                                    })

                                if new_history_rows:
                                    df_history = pd.concat([df_history, pd.DataFrame(new_history_rows)], ignore_index=True)
                                    save_data(df_history)

                                st.session_state.inventory_data_cache = workbook_data
                                save_inventory_workbook(workbook_data)

                                if st.session_state.shift_active:
                                    st.session_state.active_shift_context["accumulated_sales_expected"] += total_belanja_confirm
                                    save_active_shift({"shift_active": True, **st.session_state.active_shift_context})

                                st.session_state.nota_confirmed = True
                                st.rerun()
                else:
                    st.success("✅ Pembayaran sudah dikonfirmasi dan stok sudah diperbarui. Silakan cetak/unduh struk di panel kanan.")
                    
                    col_trx1, col_trx2 = st.columns([1, 1])
                    with col_trx1:
                        if st.button("🆕 Transaksi Baru", type="primary", use_container_width=True):
                            st.session_state.cart = []
                            st.session_state.checkout_mode = False
                            st.session_state.bayar_tunai = 0
                            st.session_state.nota_confirmed = False
                            st.rerun()
                    with col_trx2:
                        if st.button("🛑 Tutup Shift", type="secondary", use_container_width=True):
                            st.session_state.cart = []
                            st.session_state.checkout_mode = False
                            st.session_state.bayar_tunai = 0
                            st.session_state.nota_confirmed = False
                            st.session_state.target_menu = "🕒 Sesi Shift"
                            st.rerun()

    with col_nota:
        st.subheader("📄 Preview Nota")

        if st.session_state.cart:
            total_belanja = sum(item["subtotal"] for item in st.session_state.cart)
            bayar_tunai = st.session_state.bayar_tunai if st.session_state.nota_confirmed else 0
            kembali = bayar_tunai - total_belanja
            tgl_today = get_wib_time().strftime("%d/%m/%Y")
            
            kasir_mapping = {"Ivonne": "A1", "Dian": "K1", "Julia": "K2"}
            nama_raw = kasir_nama_nota if 'kasir_nama_nota' in locals() else st.session_state.active_shift_context.get("user_name", "")
            kasir_nama_nota_html = kasir_mapping.get(nama_raw, nama_raw)

            def format_angka(val):
                try: return f"{int(val):,}".replace(",", ".")
                except: return str(val)

            items_html = ""
            for item in st.session_state.cart:
                items_html += f"<div style='display: flex; justify-content: space-between; margin-bottom: 4px;'><span style='flex: 2; text-align: left;'>{item['qty']} {item['nama']}</span><span style='flex: 1; text-align: center;'>{format_angka(item['harga_per_satuan'])}</span><span style='flex: 1; text-align: right;'>{format_angka(item['subtotal'])}</span></div>"

            nota_html = f"""
<!DOCTYPE html><html><head><style>body {{ margin: 0; padding: 5px; background-color: transparent; }} * {{ box-sizing: border-box; }}</style></head>
<body><div style="font-family: 'Courier New', Courier, monospace; font-size: 11px; border: 1px solid #e0e0e0; padding: 10px; border-radius: 6px; max-width: 280px; margin: 0 auto; background-color: #f8f9fa; color: #333; box-shadow: 0px 2px 6px rgba(0,0,0,0.08);">
    <div style="text-align: center; border-bottom: 1px dashed #666; padding-bottom: 8px; margin-bottom: 8px; line-height: 1.3;">
        <b style="font-size: 13px; color: #222;">APOTEK VETERAN SEHAT BLITAR</b><br>Jl. Veteran no 64B Blitar Kota<br>(Sebelah Gang Srigading)<br>Blitar 66111<br><b>081331808585</b><br>Harga Sudah Termasuk PPN</div>
    <div style="margin-bottom: 8px; font-size: 10px; color: #555; text-align: left; word-break: break-word;">{tgl_today} <span id="clock_kasir_realtime"></span> {kasir_nama_nota_html}</div>
    <div style="border-bottom: 1px dashed #666; margin-bottom: 8px;"></div>
    <div style="font-size: 11px; word-break: break-word;">{items_html}</div>
    <div style="border-top: 1px dashed #666; margin-top: 8px; padding-top: 8px; font-size: 11px;">
        <div style='display: flex; justify-content: space-between; margin-bottom: 3px;'><b style="font-size: 12px; color: #222;">Total</b> <b style="font-size: 12px; color: #e94560;">{format_angka(total_belanja)}</b></div>
        <div style='display: flex; justify-content: space-between; margin-bottom: 3px; color: #444;'>Bayar <span>{format_angka(bayar_tunai)}</span></div>
        <div style='display: flex; justify-content: space-between; color: #444;'>Kembali <span>{format_angka(max(0, kembali))}</span></div>
    </div>
    <div style="text-align: center; margin-top: 12px; font-size: 9.5px; color: #777; line-height: 1.3;">- Terimakasih Semoga Lekas Sembuh -</div>
</div>
<script>
function updateClock() {{
    var d = new Date(); 
    var h = String(d.getHours()).padStart(2, '0'); var m = String(d.getMinutes()).padStart(2, '0'); var s = String(d.getSeconds()).padStart(2, '0');
    var el = document.getElementById('clock_kasir_realtime'); if (el) {{ el.innerHTML = h + ":" + m + ":" + s; }}
}} setInterval(updateClock, 1000); updateClock();
</script></body></html>
"""
            components.html(nota_html, height=450, scrolling=True)

            st.markdown("<br>", unsafe_allow_html=True)
            
            html_printable_nota = f"""
<!DOCTYPE html>
<html>
<head>
<title>Cetak Struk Nota - Apotek Veteran Blitar</title>
<style>
    @page {{ size: 80mm auto; margin: 0mm; }}
    * {{ box-sizing: border-box; }}
    body {{ font-family: 'Courier New', Courier, monospace; font-size: 11px; line-height: 1.3; margin: 0; padding: 4mm; color: #000; background: #fff; }}
    .print-container {{ width: 100%; max-width: 72mm; margin: 0 auto; }}
    .text-center {{ text-align: center; }}
    .header-title {{ font-size: 13px; font-weight: bold; }}
    .border-dash {{ border-bottom: 1px dashed #000; margin: 6px 0; }}
    .flex-between {{ display: flex; justify-content: space-between; margin-bottom: 2px; }}
    .info-meta {{ font-size: 10px; word-break: break-word; }}
    .items-wrapper {{ font-size: 11px; word-break: break-word; }}
    .footer-text {{ font-size: 9.5px; line-height: 1.3; }}
    .btn-container {{ text-align: center; margin-top: 15px; }}
    @media print {{ body {{ padding: 2mm; }} .btn-container {{ display: none !important; }} }}
</style>
</head>
<body>
<div class="print-container">
    <div class="text-center">
        <span class="header-title">APOTEK VETERAN SEHAT BLITAR</span><br>
        Jl. Veteran no 64B Blitar Kota<br>(Sebelah Gang Srigading)<br>081331808585<br>Harga Sudah Termasuk PPN<br>
    </div>
    <div class="border-dash"></div>
    <div class="info-meta">{tgl_today} <span id="clock_print_realtime"></span> {kasir_nama_nota_html}</div>
    <div class="border-dash"></div>
    <div class="items-wrapper">{items_html}</div>
    <div class="border-dash"></div>
    <div class="flex-between"><b>Total</b> <b>{format_rupiah(total_belanja)}</b></div>
    <div class="flex-between">Bayar <span>{format_rupiah(bayar_tunai)}</span></div>
    <div class="flex-between">Kembali <span>{format_rupiah(max(0, kembali))}</span></div>
    <div class="border-dash"></div>
    <div class="text-center footer-text">- Terimakasih Semoga Lekas Sembuh -</div>
</div>
<script>
function updatePrintClock() {{
    var d = new Date(); var h = String(d.getHours()).padStart(2, '0'); var m = String(d.getMinutes()).padStart(2, '0'); var s = String(d.getSeconds()).padStart(2, '0');
    var el = document.getElementById('clock_print_realtime'); if (el) {{ el.innerHTML = h + ":" + m + ":" + s; }}
}} setInterval(updatePrintClock, 1000); updatePrintClock();
</script></body></html>
"""
            
            b64_html = base64.b64encode(html_printable_nota.encode("utf-8")).decode("utf-8")
            custom_print_button = f"""
            <!DOCTYPE html>
            <html>
            <head>
            <style>
                body {{ margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: transparent; }}
                .btn {{ display: flex; align-items: center; justify-content: center; width: 100%; height: 40px; background-color: #ff4b4b; color: white; border: none; border-radius: 8px; font-size: 14px; font-weight: 600; cursor: pointer; transition: 0.3s; }}
                .btn:hover {{ background-color: #ff3333; }}
            </style>
            </head>
            <body>
                <button class="btn" onclick="printReceipt()">🖨️ Cetak & Print Nota</button>
                <script>
                function printReceipt() {{
                    const b64 = "{b64_html}";
                    const binStr = atob(b64);
                    const len = binStr.length;
                    const bytes = new Uint8Array(len);
                    for (let i = 0; i < len; i++) {{ bytes[i] = binStr.charCodeAt(i); }}
                    const htmlContent = new TextDecoder('utf-8').decode(bytes);
                    const printWin = window.open('', '_blank', 'width=400,height=600');
                    printWin.document.open(); printWin.document.write(htmlContent); printWin.document.close();
                    setTimeout(function() {{ printWin.focus(); printWin.print(); }}, 500);
                }}
                </script>
            </body>
            </html>
            """
            components.html(custom_print_button, height=45)

            if not st.session_state.nota_confirmed:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("🗑️ Batalkan & Kosongkan", type="secondary", use_container_width=True):
                    st.session_state.cart = []; st.session_state.checkout_mode = False; st.session_state.bayar_tunai = 0; st.session_state.nota_confirmed = False; st.rerun()
        else:
            st.info("Keranjang masih kosong. Tambahkan obat dari form di sebelah kiri.")

# ══════════════════════════════════════════════════════════════════════════════
# RETUR & ENTRY PEMBELIAN
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "📦 Retur & Entry Pembelian":
    st.title("📦 Retur & Entry Pembelian")
    st.caption("Kelola pengembalian obat dan pencatatan restok (pembelian) secara langsung ke sistem.")

    def get_dataset_options(df_current=None):
        df_inv = build_inventory_print_dataframe()
        prods, sats, batches = [], [], []
        if df_inv is not None and not df_inv.empty:
            prods = [str(x).strip() for x in df_inv["Nama produk"].dropna().unique() if str(x).strip()]
            sats = [str(x).strip() for x in df_inv["Satuan"].dropna().unique() if str(x).strip()]
            batches = [str(x).strip() for x in df_inv["Nomor Batch"].dropna().unique() if str(x).strip()]
        if df_current is not None and not df_current.empty:
            if "Nama produk" in df_current.columns: prods += [str(x).strip() for x in df_current["Nama produk"].dropna().unique() if str(x).strip()]
            if "Satuan" in df_current.columns: sats += [str(x).strip() for x in df_current["Satuan"].dropna().unique() if str(x).strip()]
            if "Nomor Batch" in df_current.columns: batches += [str(x).strip() for x in df_current["Nomor Batch"].dropna().unique() if str(x).strip()]
        return sorted(list(set(prods))), sorted(list(set(sats))), sorted(list(set(batches)))

    tab_retur, tab_entri = st.tabs(["🏥 Retur Pembelian", "🛍️ Entry Pembelian"])

    with tab_retur:
        st.markdown("<div class='app-header'><div class='app-title'>🏥 Retur Pembelian Obat</div><div class='app-subtitle'>Pilih produk dari worksheet yang sudah diupload, lalu buat retur sesuai stok real-time.</div></div>", unsafe_allow_html=True)

        if "inventory_data_cache" not in st.session_state or not st.session_state.inventory_data_cache:
            st.warning("Dataset belum tersedia. Silakan upload dataset terlebih dahulu di menu **📋 Kelola Stok**.")
            st.stop()

        workbook_data = st.session_state.inventory_data_cache
        AVAILABLE_SHEETS = get_available_sheets()
        sheet_name = st.selectbox("Pilih Worksheet", AVAILABLE_SHEETS, index=0, key="retur_selected_sheet")
        if sheet_name not in workbook_data: st.stop()

        sheet_df = prepare_sheet_for_editor(workbook_data[sheet_name].copy()).sort_values(["Nama produk", "Nomor Batch"], na_position="last").reset_index(drop=True)
        col_meta_a, col_meta_b, col_meta_c = st.columns(3)
        with col_meta_a: st.metric("Worksheet Aktif", sheet_name)
        with col_meta_b: st.metric("Jumlah Baris", len(sheet_df))
        with col_meta_c: st.metric("Total Stok Sisa", int(sheet_df["Stok Sisa"].fillna(0).sum()))

        st.markdown("---")
        search_text = st.text_input("🔍 Cari Data Retur (Nama Produk, Batch, dll)", placeholder="Ketik kata kunci pencarian...", key="retur_search_input")
        if search_text.strip():
            mask = sheet_df.astype(str).apply(lambda col: col.str.contains(search_text.strip(), case=False, na=False)).any(axis=1)
            filtered_df = sheet_df[mask].copy()
        else: filtered_df = sheet_df.copy()

        if filtered_df.empty: st.stop()

        st.subheader("📦 Pilih Produk untuk Retur")
        preview_df = filtered_df[["Nama produk", "Nomor Batch", "Satuan", "Tanggal Kadaluwarsa", "Stok Sisa", "Harga 1", "Keterangan"]].copy()
        preview_df["Tanggal Kadaluwarsa"] = preview_df["Tanggal Kadaluwarsa"].apply(lambda x: x.strftime("%d-%m-%Y") if pd.notna(x) else "")
        st.dataframe(preview_df, use_container_width=True, hide_index=True, height=260)

        product_options = filtered_df["Nama produk"].fillna("").astype(str).drop_duplicates().tolist()
        selected_product = st.selectbox("Pilih Produk", product_options, key="retur_product_select")
        
        if selected_product:
            product_rows = filtered_df[filtered_df["Nama produk"].fillna("").astype(str).str.lower() == str(selected_product).lower()].copy()
            if not product_rows.empty:
                selected_batch = st.selectbox("Pilih Nomor Batch", product_rows["Nomor Batch"].fillna("-").astype(str).drop_duplicates().tolist(), key="retur_batch_select")
                match_batch = product_rows[product_rows["Nomor Batch"].fillna("-").astype(str) == str(selected_batch)]
                if not match_batch.empty: selected_row = match_batch.iloc[0]
                else: selected_row = product_rows.iloc[0]

                with st.form("form_retur_entry"):
                    qty_retur = st.number_input("Jumlah Retur (unit)", min_value=0.0, step=1.0, value=0.0)
                    ket_val = str(selected_row.get("Keterangan", "")) if pd.notna(selected_row.get("Keterangan")) else ""
                    keterangan_retur = st.text_area("Keterangan Retur", value=ket_val, height=90)
                    if st.form_submit_button("➕ Tambahkan ke Daftar Retur", type="primary"):
                        if qty_retur <= 0: st.warning("Jumlah retur harus lebih dari 0.")
                        else:
                            exp_date = parse_excel_date(selected_row.get("Tanggal Kadaluwarsa"))
                            if pd.isna(exp_date): exp_date = get_wib_time().date()
                            else: exp_date = exp_date.date()

                            new_item = {
                                "Nama produk": selected_row.get("Nama produk", ""), "Satuan": selected_row.get("Satuan", ""), "Nomor Batch": selected_batch,
                                "Tanggal Kadaluwarsa": exp_date, "Stok Sisa": float(selected_row.get("Stok Sisa", 0)), "Jumlah Retur": float(qty_retur),
                                "Harga 1": float(selected_row.get("Harga 1", 0)), "Keterangan": keterangan_retur
                            }
                            st.session_state.retur_items = pd.concat([st.session_state.retur_items, pd.DataFrame([new_item])], ignore_index=True)
                            st.toast(f"✅ Produk **{selected_product}** berhasil ditambahkan ke daftar retur.")
                            time.sleep(1)
                            st.rerun()

        st.markdown("---")
        st.subheader("🧾 Daftar Item Retur")
        all_items_df = build_inventory_print_dataframe()

        if st.session_state.retur_items.empty:
            st.info("Belum ada item retur.")
            edited_df = st.session_state.retur_items
        else:
            opsi_produk_r, opsi_satuan_r, opsi_batch_r = get_dataset_options(st.session_state.retur_items)
            edited_df = st.data_editor(st.session_state.retur_items.copy(), use_container_width=True, num_rows="dynamic", hide_index=True,
                column_config={"Nama produk": st.column_config.SelectboxColumn("Nama Produk", options=opsi_produk_r, width="large"), "Tanggal Kadaluwarsa": st.column_config.DateColumn("Tanggal Kadaluwarsa", format="YYYY-MM-DD")}, key="data_editor_retur")
            
            for i, row in edited_df.iterrows():
                if pd.isna(row.get("Tanggal Kadaluwarsa")): edited_df.at[i, "Tanggal Kadaluwarsa"] = None
                elif isinstance(row["Tanggal Kadaluwarsa"], pd.Timestamp): edited_df.at[i, "Tanggal Kadaluwarsa"] = row["Tanggal Kadaluwarsa"].date()

            changed_retur = False
            for i, row in edited_df.iterrows():
                new_nama = str(row["Nama produk"]).strip() if pd.notna(row["Nama produk"]) else ""
                old_nama = str(st.session_state.retur_items.loc[i, "Nama produk"]).strip() if i in st.session_state.retur_items.index else ""
                if new_nama and new_nama.lower() not in ["none", "nan"] and new_nama != old_nama:
                    match = all_items_df[all_items_df["Nama produk"].astype(str).str.strip() == new_nama]
                    if not match.empty:
                        prod = match.iloc[0]
                        edited_df.at[i, "Satuan"] = str(prod["Satuan"]) if pd.notna(prod["Satuan"]) else ""
                        edited_df.at[i, "Nomor Batch"] = str(prod["Nomor Batch"]) if pd.notna(prod["Nomor Batch"]) else ""
                        t_exp = parse_excel_date(prod["Tanggal Kadaluwarsa"])
                        if pd.notna(t_exp): edited_df.at[i, "Tanggal Kadaluwarsa"] = t_exp.date()
                        edited_df.at[i, "Stok Sisa"] = float(prod["Stok Sisa"]) if pd.notna(prod["Stok Sisa"]) else 0.0
                        edited_df.at[i, "Harga 1"] = float(prod["Harga 1"]) if pd.notna(prod["Harga 1"]) else 0.0
                        changed_retur = True

            if changed_retur: st.session_state.retur_items = edited_df; st.rerun()
            else: st.session_state.retur_items = edited_df

        total_retur = float((edited_df["Jumlah Retur"].fillna(0) * edited_df["Harga 1"].fillna(0)).sum()) if not edited_df.empty else 0.0

        col_save, col_reset = st.columns([1, 1])
        with col_save:
            if st.button("💾 Simpan Retur ke Database", type="primary", use_container_width=True):
                if edited_df.empty or edited_df["Jumlah Retur"].fillna(0).sum() <= 0: st.warning("Daftar retur masih kosong atau belum ada jumlah retur yang valid.")
                else:
                    df_history = load_data()
                    new_history_rows = []
                    active_df = prepare_sheet_for_editor(workbook_data[sheet_name].copy())
                    current_petugas = USERS.get(st.session_state.username, {}).get("name", "Sistem")
                    
                    for _, item in edited_df.iterrows():
                        qty_retur_item = float(item["Jumlah Retur"]) if pd.notna(item["Jumlah Retur"]) else 0.0
                        if qty_retur_item <= 0: continue
                        nama_item = str(item["Nama produk"]).strip() if pd.notna(item["Nama produk"]) else ""
                        batch_item = str(item["Nomor Batch"]).strip() if pd.notna(item["Nomor Batch"]) else ""
                        
                        mask = ((active_df["Nama produk"].fillna("").astype(str).str.lower() == nama_item.lower()) & (active_df["Nomor Batch"].fillna("").astype(str).str.lower() == batch_item.lower()))
                        if not mask.any(): continue
                            
                        idx = active_df[mask].index[-1]
                        stok_sisa_lama = float(active_df.loc[idx, "Stok Sisa"] if pd.notna(active_df.loc[idx, "Stok Sisa"]) else 0)
                        stok_baru = max(stok_sisa_lama - qty_retur_item, 0)
                        active_df.loc[idx, "Stok Sisa"] = stok_baru
                        active_df.loc[idx, "Stok Keluar"] = float(active_df.loc[idx, "Stok Keluar"] if pd.notna(active_df.loc[idx, "Stok Keluar"]) else 0) + qty_retur_item
                        active_df.loc[idx, "Keterangan"] = str(item["Keterangan"]) if pd.notna(item["Keterangan"]) else active_df.loc[idx, "Keterangan"]

                        harga_1_item = float(item["Harga 1"]) if pd.notna(item["Harga 1"]) else 0.0
                        t_exp_s = parse_excel_date(item["Tanggal Kadaluwarsa"])
                        
                        new_history_rows.append({
                            "Tanggal": get_wib_time().strftime("%Y-%m-%d %H:%M:%S"), 
                            "Nomor Faktur": "RETUR", 
                            "Nama Obat": nama_item, 
                            "Kategori": sheet_name, 
                            "Satuan": str(item["Satuan"]) if pd.notna(item["Satuan"]) else "", 
                            "Nomor Batch": batch_item,
                            "Stok Masuk": 0.0, "Stok Keluar": qty_retur_item, "Stok Akhir": stok_baru, 
                            "Harga Satuan (Rp)": harga_1_item,
                            "Total Nilai (Rp)": qty_retur_item * harga_1_item, 
                            "Tanggal Kadaluarsa": t_exp_s if pd.notna(t_exp_s) else pd.NaT,
                            "Keterangan": f"Retur Pembelian Obat",
                            "Petugas": current_petugas
                        })

                    workbook_data[sheet_name] = active_df
                    st.session_state.inventory_data_cache = workbook_data
                    save_inventory_workbook(workbook_data)

                    if new_history_rows:
                        df_history = pd.concat([df_history, pd.DataFrame(new_history_rows)], ignore_index=True)
                        save_data(df_history)

                    history_row = pd.DataFrame([{"Nomor Faktur": str(selected_batch) if 'selected_batch' in locals() else "-", "Tanggal Retur": pd.Timestamp(get_wib_time().date()), "Jumlah Item": int(len(edited_df[edited_df["Jumlah Retur"].fillna(0) > 0])), "Total Nilai Retur": total_retur, "Tanggal Disimpan": get_wib_time()}])
                    st.session_state.retur_history = pd.concat([st.session_state.retur_history, history_row], ignore_index=True)
                    save_retur_history(st.session_state.retur_history)

                    st.toast("✅ Retur berhasil disimpan!")
                    time.sleep(1)
                    st.session_state.retur_items = pd.DataFrame(columns=st.session_state.retur_items.columns)
                    st.rerun()
        with col_reset:
            if st.button("🔄 Reset Daftar Retur", type="secondary", use_container_width=True):
                st.session_state.retur_items = pd.DataFrame(columns=st.session_state.retur_items.columns); st.rerun()

        st.markdown("---")
        st.subheader("📜 Riwayat Retur")
        if st.session_state.retur_history.empty: st.info("Belum ada riwayat retur.")
        else:
            history_display = st.session_state.retur_history.copy()
            history_display["Tanggal Retur"] = pd.to_datetime(history_display["Tanggal Retur"]).apply(lambda x: x.strftime("%d-%m-%Y") if pd.notna(x) else "")
            history_display["Tanggal Disimpan"] = pd.to_datetime(history_display["Tanggal Disimpan"]).apply(lambda x: x.strftime("%d-%m-%Y %H:%M:%S") if pd.notna(x) else "")
            history_display["Total Nilai Retur"] = pd.to_numeric(history_display["Total Nilai Retur"], errors="coerce").fillna(0).apply(lambda x: f"Rp {x:,.2f}".replace(",", "."))
            st.dataframe(history_display, use_container_width=True, hide_index=True)

    with tab_entri:
        st.markdown("<div class='app-header'><div class='app-title'>🛍️ Entry Pembelian Obat</div><div class='app-subtitle'>Catat restok secara ringkas, dan simpan langsung ke Database.</div></div>", unsafe_allow_html=True)
        if "inventory_data_cache" not in st.session_state or not st.session_state.inventory_data_cache: st.stop()
            
        st.caption("Pencarian obat dilakukan dari seluruh worksheet. Entry pembelian ini akan langsung menambah riwayat pada Database.")
        no_faktur = st.text_input("No. Faktur Pembelian", key="no_faktur_pembelian")
        pbf_default = st.text_input("PBF (Distributor) Default", key="pbf_pembelian")
        
        all_items_df = build_inventory_print_dataframe()
        cari_obat_input = st.text_input("🔍 Pencarian Produk (Semua Data: Nama, Batch, Faktur, dll)", placeholder="Ketik kata kunci pencarian...", key="cari_obat_pembelian_input")
        
        if cari_obat_input.strip() and all_items_df is not None and not all_items_df.empty:
            mask = all_items_df.astype(str).apply(lambda col: col.str.contains(cari_obat_input.strip(), case=False, na=False)).any(axis=1)
            hasil = all_items_df[mask]
            
            if not hasil.empty:
                st.success(f"Ditemukan {len(hasil)} entri. Pilih salah satu baris di bawah, lalu klik Tambahkan:")
                tabel_cari_df = hasil[["Worksheet", "Nama produk", "Satuan", "Harga 1", "Harga 2", "Stok Sisa"]].drop_duplicates(subset=["Worksheet", "Nama produk"]).reset_index(drop=True)
                event_beli = st.dataframe(tabel_cari_df, use_container_width=True, hide_index=True, on_select="rerun", selection_mode="single-row", key="table_hasil_pencarian_pembelian")
                
                if event_beli.selection.rows:
                    idx = event_beli.selection.rows[0]
                    if idx < len(tabel_cari_df):
                        selected_row = tabel_cari_df.iloc[idx]
                        if st.button(f"➕ Tambahkan '{selected_row['Nama produk']}' ke Tabel Entry", key="tambah_ke_pembelian"):
                            new_row = {
                                "No.": len(st.session_state.df_beli) + 1, "Worksheet": selected_row["Worksheet"], "Nama produk": selected_row["Nama produk"],
                                "Satuan": selected_row["Satuan"], "Nomor Batch": "", "Tanggal Kadaluwarsa": (get_wib_time() + timedelta(days=365)).date(),
                                "Stok Masuk": 0.0, "Harga 1": float(selected_row["Harga 1"]) if pd.notna(selected_row["Harga 1"]) else 0.0,
                                "Harga 2": float(selected_row["Harga 2"]) if pd.notna(selected_row["Harga 2"]) else 0.0, "Keterangan": ""
                            }
                            df_existing = st.session_state.df_beli
                            if len(df_existing) == 1 and not str(df_existing.iloc[0]["Nama produk"]).strip(): st.session_state.df_beli = pd.DataFrame([new_row])
                            else: st.session_state.df_beli = pd.concat([df_existing, pd.DataFrame([new_row])], ignore_index=True)
                            st.toast(f"✅ {selected_row['Nama produk']} ditambahkan ke tabel entry!")
                            time.sleep(0.5)
                            st.rerun()

        st.markdown("---")
        st.subheader("📦 Rincian Item Entry")
        if "df_beli" not in st.session_state:
            st.session_state.df_beli = pd.DataFrame([{"No.": 1, "Worksheet": "TAB", "Nama produk": "", "Satuan": "TAB", "Nomor Batch": "", "Tanggal Kadaluwarsa": get_wib_time().date(), "Stok Masuk": 0.0, "Harga 1": 0.0, "Harga 2": 0.0, "Keterangan": ""}])
            
        opsi_produk_e, _, _ = get_dataset_options(st.session_state.df_beli)
        AVAILABLE_SHEETS = get_available_sheets()

        edited_df = st.data_editor(
            st.session_state.df_beli.copy(), use_container_width=True, num_rows="dynamic", hide_index=True,
            column_config={
                "Worksheet": st.column_config.SelectboxColumn("Worksheet Tujuan", options=AVAILABLE_SHEETS, width="small"),
                "Nama produk": st.column_config.SelectboxColumn("Nama Produk", options=opsi_produk_e, width="large"),
                "Tanggal Kadaluwarsa": st.column_config.DateColumn("Exp Date", format="YYYY-MM-DD")
            }, key="df_beli_editor"
        )
        
        for i, row in edited_df.iterrows():
            if pd.isna(row.get("Tanggal Kadaluwarsa")): edited_df.at[i, "Tanggal Kadaluwarsa"] = None
            elif isinstance(row["Tanggal Kadaluwarsa"], pd.Timestamp): edited_df.at[i, "Tanggal Kadaluwarsa"] = row["Tanggal Kadaluwarsa"].date()

        changed_beli = False
        for i, row in edited_df.iterrows():
            new_nama = str(row["Nama produk"]).strip() if pd.notna(row.get("Nama produk")) else ""
            old_nama = str(st.session_state.df_beli.loc[i, "Nama produk"]).strip() if i in st.session_state.df_beli.index else ""
            if new_nama and new_nama.lower() not in ["none", "nan"] and new_nama != old_nama:
                match = all_items_df[all_items_df["Nama produk"].astype(str).str.strip() == new_nama]
                if not match.empty:
                    prod = match.iloc[0]
                    edited_df.at[i, "Worksheet"] = prod["Worksheet"]
                    edited_df.at[i, "Satuan"] = str(prod["Satuan"]) if pd.notna(prod["Satuan"]) else ""
                    t_exp = parse_excel_date(prod["Tanggal Kadaluwarsa"])
                    if pd.notna(t_exp): edited_df.at[i, "Tanggal Kadaluwarsa"] = t_exp.date()
                    edited_df.at[i, "Harga 1"] = float(prod["Harga 1"]) if pd.notna(prod["Harga 1"]) else 0.0
                    edited_df.at[i, "Harga 2"] = float(prod["Harga 2"]) if pd.notna(prod["Harga 2"]) else 0.0
                    changed_beli = True

        if changed_beli: st.session_state.df_beli = edited_df; st.rerun()
        else: st.session_state.df_beli = edited_df
        
        col_simpan_beli, col_reset_beli = st.columns([1, 1])
        with col_simpan_beli:
            if st.button("💾 Simpan Entry ke Database", type="primary", use_container_width=True):
                has_valid_item = any(pd.notna(r["Nama produk"]) and str(r["Nama produk"]).strip() != "" and str(r["Nama produk"]).strip().lower() != "none" for _, r in edited_df.iterrows())
                if edited_df.empty or not has_valid_item: st.warning("Tabel entry kosong atau nama produk belum diisi secara valid.")
                else:
                    workbook_data = st.session_state.inventory_data_cache
                    jumlah_disimpan = 0
                    df_history = load_data()
                    new_history_rows = []
                    current_petugas = USERS.get(st.session_state.username, {}).get("name", "Sistem")
                    
                    for _, row in edited_df.iterrows():
                        nama = str(row["Nama produk"]).strip() if pd.notna(row["Nama produk"]) else ""
                        stok_masuk = float(row["Stok Masuk"]) if pd.notna(row["Stok Masuk"]) else 0.0
                        ws_target = str(row["Worksheet"]) if pd.notna(row["Worksheet"]) else ""
                        
                        if not nama or nama.lower() in ["none", "nan"] or stok_masuk <= 0 or ws_target not in workbook_data: continue
                            
                        sheet_df = prepare_sheet_for_editor(workbook_data[ws_target].copy())
                        harga1_beli = float(row["Harga 1"]) if pd.notna(row["Harga 1"]) else 0.0
                        
                        t_exp_e = parse_excel_date(row["Tanggal Kadaluwarsa"])
                        tgl_exp = t_exp_e if pd.notna(t_exp_e) else pd.Timestamp(get_wib_time().date() + timedelta(days=365))

                        new_buy = {
                            "Nama produk": nama, "Satuan": str(row["Satuan"]) if pd.notna(row["Satuan"]) else "",
                            "Tanggal": pd.Timestamp(get_wib_time().date()), "Nomor Faktur": no_faktur, "Nomor Batch": str(row["Nomor Batch"]) if pd.notna(row["Nomor Batch"]) else "",
                            "PBF": pbf_default, "Tanggal Kadaluwarsa": tgl_exp, "Stok Masuk": stok_masuk, "Stok Keluar": 0.0, "Stok Sisa": stok_masuk,
                            "Harga 1": harga1_beli, "Harga 2": float(row["Harga 2"]) if pd.notna(row["Harga 2"]) else 0.0, "Keterangan": str(row["Keterangan"]) if pd.notna(row["Keterangan"]) else ""
                        }
                        
                        sheet_df = pd.concat([sheet_df, pd.DataFrame([new_buy])], ignore_index=True)
                        workbook_data[ws_target] = sheet_df
                        jumlah_disimpan += 1
                        
                        new_history_rows.append({
                            "Tanggal": get_wib_time().strftime("%Y-%m-%d %H:%M:%S"), 
                            "Nomor Faktur": no_faktur,
                            "Nama Obat": nama, "Kategori": ws_target, "Satuan": new_buy["Satuan"], "Nomor Batch": str(row["Nomor Batch"]) if pd.notna(row["Nomor Batch"]) else "",
                            "Stok Masuk": stok_masuk, "Stok Keluar": 0.0, "Stok Akhir": stok_masuk, 
                            "Harga Satuan (Rp)": harga1_beli, "Total Nilai (Rp)": stok_masuk * harga1_beli,
                            "Tanggal Kadaluarsa": tgl_exp, "Keterangan": f"Pembelian/Stok Masuk Obat",
                            "Petugas": current_petugas
                        })
                        
                    if jumlah_disimpan > 0:
                        st.session_state.inventory_data_cache = workbook_data
                        save_inventory_workbook(workbook_data)
                        if new_history_rows: save_data(pd.concat([df_history, pd.DataFrame(new_history_rows)], ignore_index=True))
                        st.session_state.df_beli = pd.DataFrame([{"No.": 1, "Worksheet": "TAB", "Nama produk": "", "Satuan": "", "Nomor Batch": "", "Tanggal Kadaluwarsa": get_wib_time().date(), "Stok Masuk": 0.0, "Harga 1": 0.0, "Harga 2": 0.0, "Keterangan": ""}])
                        st.toast(f"✅ {jumlah_disimpan} entri berhasil disimpan langsung ke Database!")
                        time.sleep(1)
                        st.rerun()
                    else: st.warning("Tidak ada item valid untuk disimpan.")

        with col_reset_beli:
            if st.button("🗑️ Reset Tabel Entry", type="secondary", use_container_width=True):
                st.session_state.df_beli = pd.DataFrame([{"No.": 1, "Worksheet": "TAB", "Nama produk": "", "Satuan": "", "Nomor Batch": "", "Tanggal Kadaluwarsa": get_wib_time().date(), "Stok Masuk": 0.0, "Harga 1": 0.0, "Harga 2": 0.0, "Keterangan": ""}])
                st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# SESI SHIFT
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "🕒 Sesi Shift":

    def format_angka_erp(val):
        try: return f"{float(val):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except: return "0,00"

    def render_row_erp(label, val_num=0.0, disabled=True, widget="text", opts=None, val_str="", key_suffix="", masked=False):
        c1, c2 = st.columns([3, 7])
        k = f"ts_{key_suffix}_{re.sub(r'[^a-zA-Z0-9]', '_', label)}"
        with c1: st.markdown(f"<div style='text-align: right; padding-top: 8px; font-weight: 600; font-size: 13px; color: #e0e0e0;'>{label}</div>", unsafe_allow_html=True)
        with c2:
            if masked:
                st.text_input(label, value="*** (Disembunyikan)", disabled=True, label_visibility="collapsed", key=k)
                return val_num
            if disabled:
                if widget == "number":
                    st.text_input(label, value=format_angka_erp(val_num), disabled=True, label_visibility="collapsed", key=k)
                    return val_num
                else:
                    st.text_input(label, value=val_str, disabled=True, label_visibility="collapsed", key=k)
                    return val_str
            else:
                if widget == "number": return st.number_input(label, value=float(val_num), label_visibility="collapsed", key=k, step=1000.0, format="%.2f")
                elif widget == "select": return st.selectbox(label, options=opts, index=opts.index(val_str) if val_str in opts else 0, label_visibility="collapsed", key=k)
                elif widget == "text": return st.text_input(label, value=val_str, label_visibility="collapsed", key=k)

    kasir_options = ["Ivonne", "Dian", "Julia"] if st.session_state.role == "Admin" else ["Dian", "Julia"]

    if st.session_state.get("step_tutup_shift") == 3 and "last_shift_data" in st.session_state:
        st.markdown("<h2 style='text-align: center; margin-bottom: 10px; color: #e0e0e0;'>Laporan Tutup Shift</h2>", unsafe_allow_html=True)
        st.success("✅ Shift berhasil ditutup. Berikut adalah laporan data Anda.")
        
        df_report = st.session_state.last_shift_data.copy()
        df_preview = df_report.copy()
        for col in ["Saldo Awal", "Hasil Penjualan", "Piutang", "Pendapatan Jurnal", "Total Pendapatan", "Retur Penjualan", "Pengeluaran Jurnal", "Total Pengeluaran", "Saldo Akhir Sistem", "Fisik Kasir Aktual", "Selisih"]:
            if col in df_preview.columns: df_preview[col] = df_preview[col].apply(lambda x: format_rupiah(x))
                
        st.markdown("---")
        st.subheader("👁️ Preview Laporan Shift")
        st.dataframe(df_preview, use_container_width=True, hide_index=True)
        
        col_d1, col_d2, col_d3, col_d4 = st.columns(4)
        csv_data = df_report.to_csv(index=False).encode("utf-8-sig")
        col_d1.download_button("📄 Unduh CSV", data=csv_data, file_name=f"Shift_{df_report['Pendaftar Shift'].iloc[0]}_{get_wib_time().date()}.csv", mime="text/csv", use_container_width=True)
        
        st.write("")
        st.markdown("---")
        if st.button("✅ Selesai & Kembali ke Dashboard", type="primary", use_container_width=True):
            st.session_state.step_tutup_shift = 1
            st.session_state.target_menu = "🏠 Dashboard"
            del st.session_state.last_shift_data
            st.rerun()

    else:
        tab_aktif, tab_riwayat = st.tabs(["🕒 Sesi Shift Saat Ini", "📜 Riwayat & Laporan Shift"])

        with tab_aktif:
            if not st.session_state.shift_active:
                st.markdown("<h2 style='text-align: center; margin-bottom: 20px; color: #e0e0e0;'>Buka Shift Baru</h2>", unsafe_allow_html=True)
                auto_shift = get_auto_shift_name()
                st.info(f"Sistem mendeteksi jam saat ini otomatis masuk sebagai **Shift {auto_shift}**.")
                st.caption("Kasir yang datang pertama bertugas membuka shift dan mengisi saldo uang tunai awal laci.")
                
                default_name = USERS.get(st.session_state.username, {}).get("name", kasir_options[0])

                with st.form("form_buka_shift"):
                    nama_user_buka = render_row_erp("Kasir Pertama (Membuka)", disabled=False, widget="select", opts=kasir_options, val_str=default_name, key_suffix="buka")
                    saldo_awal_buka = render_row_erp("Saldo Tunai Awal (Modal Laci)", val_num=0.0, disabled=False, widget="number", key_suffix="buka")

                    st.write("")
                    c_btn1, c_btn2 = st.columns([3, 7])
                    with c_btn2:
                        submit_buka = st.form_submit_button("✔ Buka Shift", type="primary", use_container_width=True)

                if submit_buka:
                    st.session_state.shift_active = True
                    st.session_state.active_shift_context["saldo_awal"] = float(saldo_awal_buka)
                    st.session_state.active_shift_context["accumulated_sales_expected"] = 0.0
                    st.session_state.active_shift_context["start_time"] = get_wib_time().strftime("%Y-%m-%d %H:%M:%S")
                    st.session_state.active_shift_context["user_name"] = nama_user_buka
                    st.session_state.active_shift_context["joined_users"] = []
                    st.session_state.active_shift_context["shift_name"] = auto_shift
                    
                    save_active_shift({"shift_active": True, **st.session_state.active_shift_context})
                    st.session_state.target_menu = "🛒 Kasir Utama"
                    st.rerun()

            else:
                st.markdown("<h2 style='text-align: center; margin-bottom: 20px; color: #e0e0e0;'>Sesi Sedang Berjalan</h2>", unsafe_allow_html=True)
                
                nama_pembuka = st.session_state.active_shift_context.get("user_name", "")
                waktu_mulai = st.session_state.active_shift_context.get("start_time", "")
                shift_name = st.session_state.active_shift_context.get("shift_name", "")
                joined_users = st.session_state.active_shift_context.get("joined_users", [])

                st.info(f"🟢 **Shift {shift_name}** sedang berjalan. Dibuka oleh **{nama_pembuka}** pada {waktu_mulai}.")
                
                current_login_name = USERS.get(st.session_state.username, {}).get("name", "Unknown")
                
                if current_login_name != nama_pembuka and current_login_name not in joined_users:
                    st.warning(f"Sistem mendeteksi Anda ({current_login_name}) login namun belum tergabung di sesi ini.")
                    if st.button("🤝 Gabung Shift Ini", type="primary"):
                        st.session_state.active_shift_context["joined_users"].append(current_login_name)
                        save_active_shift({"shift_active": True, **st.session_state.active_shift_context})
                        st.success("Anda berhasil bergabung ke shift ini. Sekarang Anda bisa mengakses Kasir Utama.")
                        st.rerun()
                elif current_login_name in joined_users:
                    st.success(f"Anda ({current_login_name}) berstatus *Joined* (Tergabung) di shift ini bersama {nama_pembuka}.")
                    
                st.markdown("---")

                saldo_awal_context = st.session_state.active_shift_context["saldo_awal"]
                penjualan_sistem = st.session_state.active_shift_context["accumulated_sales_expected"]
                total_pendapatan_calc = saldo_awal_context + penjualan_sistem
                saldo_akhir_calc = total_pendapatan_calc

                if st.session_state.step_tutup_shift == 1:
                    st.markdown("### Tutup Shift")
                    st.info("💡 Langkah 1: Masukkan jumlah total uang tunai yang ada di laci kasir saat ini. Total sistem disembunyikan agar perhitungan fisik akurat.")
                    with st.form("form_input_fisik"):
                        saldo_kasir_in = st.number_input("Total Fisik Laci Kasir (Rp)", min_value=0.0, step=1000.0, value=0.0)
                        st.write("")
                        submit_fisik = st.form_submit_button("Hitung & Lanjutkan ➡️", type="primary")
                        if submit_fisik:
                            st.session_state.input_saldo_kasir = saldo_kasir_in
                            st.session_state.step_tutup_shift = 2
                            st.rerun()

                elif st.session_state.step_tutup_shift == 2:
                    st.markdown("### Verifikasi Tutup Shift")
                    saldo_kasir_in = st.session_state.input_saldo_kasir
                    selisih_calc = saldo_kasir_in - saldo_akhir_calc

                    blind_mode = False
                    if st.session_state.role != "Admin":
                        blind_mode = True
                        st.info("🔒 Mode Kasir: Nilai ekspektasi sistem disembunyikan (Blind Close). Selisih akan dikalkulasi di latar belakang untuk Admin.")
                    else:
                        st.info("💡 Mode Admin: Anda dapat memilih untuk menyembunyikan atau melihat detail sistem saat ini.")
                        blind_mode = st.checkbox("Sembunyikan Nilai Sistem (Blind Close)", value=False)

                    render_row_erp("Saldo Awal (Modal)", val_num=saldo_awal_context, disabled=True, widget="number", key_suffix="ts_awal", masked=blind_mode)
                    render_row_erp("Akumulasi Penjualan", val_num=penjualan_sistem, disabled=True, widget="number", key_suffix="ts_jual", masked=blind_mode)
                    render_row_erp("Ekspektasi Saldo Akhir Sistem", val_num=saldo_akhir_calc, disabled=True, widget="number", key_suffix="ts_akhir", masked=blind_mode)
                    
                    st.markdown("<br>", unsafe_allow_html=True)
                    render_row_erp("Saldo Kasir (Inputan Fisik Anda)", val_num=saldo_kasir_in, disabled=True, widget="number", key_suffix="ts_kasir_lock", masked=False)
                    
                    st.markdown("---")
                    
                    if not blind_mode:
                        render_row_erp("Selisih (Fisik - Sistem)", val_num=selisih_calc, disabled=True, widget="number", key_suffix="ts_selisih")
                        if selisih_calc < 0: st.error(f"⚠️ Terdapat selisih Minus: {format_rupiah(selisih_calc)}.")
                        elif selisih_calc > 0: st.warning(f"⚠️ Terdapat selisih Plus: {format_rupiah(selisih_calc)}.")
                        else: st.success("✅ Saldo Balance (Sesuai).")

                    diserahkan_kepada_opsi = ["Ivonne", "Dian", "Julia"]
                    diserahkan_kepada = render_row_erp("Shift Selanjutnya/Diserahkan Ke", disabled=False, widget="select", opts=diserahkan_kepada_opsi, key_suffix="ts_serah")
                    catatan = render_row_erp("Catatan Khusus", disabled=False, widget="text", key_suffix="ts_catatan")

                    st.write("")
                    c_space, c_btn = st.columns([6, 4])
                    with c_btn:
                        submit_tutup = st.button("✔ Konfirmasi Tutup Shift", type="primary", use_container_width=True)

                    if submit_tutup:
                        if selisih_calc != 0 and str(catatan).strip() == "":
                            st.error("❌ Karena terdapat selisih uang, wajib memberikan catatan (contoh: untuk uang parkir, selisih kembalian, dll).")
                        else:
                            log_df = load_shift_log()
                            waktu_tutup_realtime = get_wib_time().strftime("%Y-%m-%d %H:%M:%S")
                            new_log = pd.DataFrame([{
                                "Waktu Buka": waktu_mulai, "Waktu Tutup": waktu_tutup_realtime, "Shift": shift_name,
                                "Pendaftar Shift": nama_pembuka, "Kasir Bergabung": ", ".join(joined_users) if joined_users else "-",
                                "Saldo Awal": saldo_awal_context, "Hasil Penjualan": penjualan_sistem, "Piutang": 0.0, "Pendapatan Jurnal": 0.0,
                                "Total Pendapatan": total_pendapatan_calc, "Retur Penjualan": 0.0, "Pengeluaran Jurnal": 0.0, "Total Pengeluaran": 0.0,
                                "Saldo Akhir Sistem": saldo_akhir_calc, "Fisik Kasir Aktual": saldo_kasir_in, "Selisih": selisih_calc,
                                "Diserahkan Ke": diserahkan_kepada, "Catatan": catatan
                            }])
                            log_df = pd.concat([log_df, new_log], ignore_index=True)
                            save_shift_log(log_df)

                            st.session_state.shift_active = False
                            st.session_state.active_shift_context = {"saldo_awal": 0.0, "accumulated_sales_expected": 0.0, "start_time": None, "user_name": "", "joined_users": [], "shift_name": get_auto_shift_name()}

                            clear_active_shift()

                            st.session_state.last_shift_data = new_log
                            st.session_state.step_tutup_shift = 3
                            st.session_state.input_saldo_kasir = 0.0
                            st.rerun()

        with tab_riwayat:
            st.markdown("### Laporan Riwayat Shift")
            log_df = load_shift_log()
            if log_df.empty:
                st.info("Belum ada riwayat shift yang tersimpan.")
            else:
                st.dataframe(log_df, use_container_width=True, hide_index=True)
                st.caption("ℹ️ Tips SOP: Laporan ini hanya untuk pengecekan kecocokan uang fisik dan sistem. Pembatalan/Retur harus dilakukan di dalam jam shift yang sama agar laporan tidak terganggu.")

# ── Footer ────────────────────────────────────────────────────────────────────
st.sidebar.markdown("---")
st.sidebar.caption("© Apotek Veteran Blitar")
